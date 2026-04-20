"""Write the Rey Capital Scalping Trading Journal xlsx from a normalised frame.

All analytics are pre-computed in transform.py and passed in as pandas
DataFrames / JournalStats — this module's only job is Excel-native rendering
(tables, conditional formatting, charts, colour bands).
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from . import transform
from .rey_theme import (
    BAND_FILL,
    BASE_FONT,
    BOLD_FONT,
    BORDER,
    CENTER,
    FONT_NAME,
    HEADER_FILL,
    HEADER_FONT,
    LEFT,
    NEGATIVE_RED,
    POSITIVE_GREEN,
    REY_BLUE,
    REY_BLUE_LIGHT,
    SUBHEADER_FILL,
    SUBHEADER_FONT,
    add_confidential_header,
    add_confidential_strip,
    add_logo,
    apply_band,
    set_widths,
    style_header,
    title_row,
)

CURRENCY_FMT = "#,##0.00;[Red]-#,##0.00"
PERCENT_FMT = "0.0%"
INT_FMT = "#,##0"
PRICE_FMT = "#,##0.00000"
DT_FMT = "yyyy-mm-dd hh:mm:ss"


# ---------------------------------------------------------------------------
# Small writing helpers
# ---------------------------------------------------------------------------

def _write_headers(ws, headers: Iterable[str], row: int) -> int:
    cols = list(headers)
    for i, h in enumerate(cols, start=1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(cols))
    return len(cols)


def _write_df(ws, df: pd.DataFrame, start_row: int, number_formats: dict | None = None) -> int:
    """Write a DataFrame's rows below an existing header row. Returns last row index."""
    if df.empty:
        return start_row - 1
    number_formats = number_formats or {}
    for r_offset, (_, row) in enumerate(df.iterrows()):
        r = start_row + r_offset
        for c_idx, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            cell = ws.cell(row=r, column=c_idx, value=val)
            if col in number_formats:
                cell.number_format = number_formats[col]
    apply_band(ws, start_row, start_row + len(df) - 1, len(df.columns))
    return start_row + len(df) - 1


def _add_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------

def _build_cover(ws, meta: dict) -> int:
    title_row(ws, "REY CAPITAL — DAILY SCALPING TRADING JOURNAL", 4)
    rows = [
        ("Generated At (UTC)", meta["generated_at"]),
        ("Data Source", meta["source"]),
        ("Accounts Included", meta["accounts"]),
        ("Date Range", meta["date_range"]),
        ("Total Trades", meta["total_trades"]),
        ("Net P&L", meta["net_profit_str"]),
        ("Currency", meta["currency"]),
        ("Version", "v1.0"),
    ]
    for i, (label, val) in enumerate(rows, start=3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.fill = SUBHEADER_FILL
        label_cell.font = SUBHEADER_FONT
        label_cell.border = BORDER
        label_cell.alignment = LEFT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        value_cell = ws.cell(row=i, column=2, value=val)
        value_cell.fill = PatternFill("solid", start_color=REY_BLUE_LIGHT)
        value_cell.font = BASE_FONT
        value_cell.alignment = LEFT
        value_cell.border = BORDER
        ws.row_dimensions[i].height = 22

    set_widths(ws, [28, 28, 20, 20])
    return len(rows) + 3


ABOUT_SECTIONS = [
    ("Purpose", [
        "This workbook is an auto-generated, read-only reflection of your closed-trade history "
        "pulled directly from MetaTrader 5. It is re-built end-to-end every time you run the "
        "refresh script — no manual data entry is required or supported.",
    ]),
    ("How to refresh", [
        "1. Ensure MT5 is installed and your accounts are configured in apps/tools/pnl_dashboard/accounts.yaml.",
        "2. From the repo root, run: python -m apps.tools.trading_journal.build_journal --config apps/tools/pnl_dashboard/accounts.yaml",
        "3. The script produces a timestamped .xlsx plus a ReyCapital_ScalpingJournal_latest.xlsx pointer in apps/tools/trading_journal/out/.",
        "Offline / CSV mode: python -m apps.tools.trading_journal.build_journal --csv <path-to-combined.csv> for rebuilding from an existing MT5 extract (useful on non-Windows machines).",
    ]),
    ("Sheets at a glance", [
        "Cover — run metadata and top-line numbers.",
        "About — this page.",
        "Trade Log — every closed trade with price, P&L, holding time, session, MFE/MAE.",
        "Daily Summary — one row per trading day with win rate, expectancy, running equity.",
        "Equity Curve — cumulative net P&L line with drawdown shading.",
        "Calendar — month grid colour-coded by daily net P&L.",
        "By Hour — P&L and win% per UTC hour (finds your productive scalping window).",
        "By Session — Asia / London / Overlap / New York split.",
        "By Symbol — per-instrument performance.",
        "By Weekday — day-of-week bias.",
        "Holding Buckets — <30s / 30s-2m / 2-5m / 5-15m / 15m+ scalping distribution.",
        "MFE/MAE — excursion efficiency versus realised move.",
        "Risk Metrics — expectancy, profit factor, max drawdown, streaks, commission drag.",
        "By Account — per-account breakdown (acts as the account filter in v1).",
    ]),
    ("Metric definitions", [
        "Win rate = wins / trades. A winning trade is any trade with net P&L > 0 after commission and swap.",
        "Expectancy = net P&L / trades. Average $ earned per trade.",
        "Profit factor = gross winning P&L / |gross losing P&L|. >1 means the system is profitable.",
        "Max drawdown = largest peak-to-trough drop in cumulative net P&L on a trade-sequenced basis.",
        "R-multiple = price move in favour / absolute stop distance. Shown blank when SL was not recorded at entry (MT5's history_deals_get currently does not surface SL on exit deals).",
        "MFE / MAE = Maximum Favorable / Adverse Excursion in price units between entry and exit, sampled from M1 bars.",
        "MFE efficiency = realised move / MFE. 1.0 means you captured the full favourable swing; 0.3 means you left 70% on the table.",
        "Session buckets use exit_time in UTC: Asia 22-07, London 07-12, Overlap 12-16, New York 16-21, Gap 21-22.",
        "Holding buckets are scalping-calibrated: <30s, 30s-2m, 2-5m, 5-15m, 15m+.",
    ]),
    ("Known limitations (v1)", [
        "Closed trades only — open positions are filtered out.",
        "No manual overlay columns (setup tag, emotion, screenshot, review note). Adding those would require append-merge logic; punted to v2.",
        "R-multiple and SL/TP-hit % depend on SL/TP being populated on exit deals; MT5's history_deals_get currently returns them as 0. Trade-level SL/TP will start showing up automatically once upstream captures them.",
        "Session classification is UTC-based. Broker server-time drift is not corrected.",
    ]),
]


def _build_about(ws) -> int:
    title_row(ws, "ABOUT THIS JOURNAL", 2)
    row = 3
    for section_title, paragraphs in ABOUT_SECTIONS:
        hdr = ws.cell(row=row, column=1, value=section_title)
        hdr.fill = SUBHEADER_FILL
        hdr.font = SUBHEADER_FONT
        hdr.alignment = LEFT
        hdr.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        for p in paragraphs:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws.cell(row=row, column=1, value=p)
            cell.font = BASE_FONT
            cell.alignment = LEFT
            cell.border = BORDER
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1  # spacer

    set_widths(ws, [28, 100])
    return row


TRADE_LOG_COLS = [
    ("account_id", "Account", 14, None),
    ("account_label", "Account Label", 18, None),
    ("ticket", "Ticket", 12, INT_FMT),
    ("symbol", "Symbol", 12, None),
    ("direction_str", "Side", 8, None),
    ("volume", "Lots", 8, "0.00"),
    ("entry_time", "Entry Time (UTC)", 20, DT_FMT),
    ("exit_time", "Exit Time (UTC)", 20, DT_FMT),
    ("holding_seconds", "Hold (s)", 10, INT_FMT),
    ("holding_bucket", "Hold Bucket", 12, None),
    ("entry_price", "Entry Px", 14, PRICE_FMT),
    ("exit_price", "Exit Px", 14, PRICE_FMT),
    ("sl", "SL", 12, PRICE_FMT),
    ("tp", "TP", 12, PRICE_FMT),
    ("pips", "Pips", 10, "0.0"),
    ("profit_loss", "Gross P&L", 14, CURRENCY_FMT),
    ("commission", "Commission", 12, CURRENCY_FMT),
    ("swap", "Swap", 10, CURRENCY_FMT),
    ("net_profit", "Net P&L", 14, CURRENCY_FMT),
    ("mfe", "MFE", 10, PRICE_FMT),
    ("mae", "MAE", 10, PRICE_FMT),
    ("mfe_efficiency", "MFE Eff.", 10, PERCENT_FMT),
    ("r_multiple", "R", 8, "0.00"),
    ("session", "Session", 16, None),
    ("hour_of_day", "Hr", 6, INT_FMT),
    ("weekday", "Day", 12, None),
    ("magic_number", "Magic", 10, INT_FMT),
    ("comment", "Comment", 24, None),
]


def _build_trade_log(ws, df: pd.DataFrame) -> int:
    title_row(ws, "TRADE LOG", len(TRADE_LOG_COLS))
    header_row = 2
    _write_headers(ws, [c[1] for c in TRADE_LOG_COLS], header_row)

    if df.empty:
        ws.cell(row=3, column=1, value="No closed trades in source data.")
        set_widths(ws, [c[2] for c in TRADE_LOG_COLS])
        return 3

    # Build a view containing only the columns the sheet expects, in the right order.
    view_cols = [src for src, _, _, _ in TRADE_LOG_COLS if src in df.columns]
    view = df[view_cols].copy()

    number_formats = {src: fmt for src, _, _, fmt in TRADE_LOG_COLS if fmt}
    last_row = _write_df(ws, view, start_row=3, number_formats=number_formats)

    # Conditional format: Net P&L column green/red.
    net_idx = view_cols.index("net_profit") + 1 if "net_profit" in view_cols else None
    if net_idx is not None and last_row >= 3:
        col_letter = get_column_letter(net_idx)
        rng = f"{col_letter}3:{col_letter}{last_row}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThan", formula=["0"],
                fill=PatternFill("solid", start_color="E2EFDA"),
                font=Font(name=FONT_NAME, size=10, color=POSITIVE_GREEN, bold=True),
            ),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="lessThan", formula=["0"],
                fill=PatternFill("solid", start_color="FCE4E4"),
                font=Font(name=FONT_NAME, size=10, color=NEGATIVE_RED, bold=True),
            ),
        )

    # Excel Table for native filter/sort.
    end_col_letter = get_column_letter(len(view_cols))
    _add_table(ws, f"A{header_row}:{end_col_letter}{last_row}", "TradeLogTable")

    set_widths(ws, [c[2] for (i, c) in enumerate(TRADE_LOG_COLS) if c[0] in view_cols])
    ws.freeze_panes = f"A{header_row + 1}"
    return last_row


DAILY_COLS = [
    ("date", "Date", 12, "yyyy-mm-dd"),
    ("trades", "Trades", 8, INT_FMT),
    ("wins", "Wins", 7, INT_FMT),
    ("losses", "Losses", 8, INT_FMT),
    ("win_rate", "Win %", 9, PERCENT_FMT),
    ("gross_profit", "Gross +", 12, CURRENCY_FMT),
    ("gross_loss", "Gross -", 12, CURRENCY_FMT),
    ("net_profit", "Net P&L", 12, CURRENCY_FMT),
    ("expectancy", "Expectancy", 12, CURRENCY_FMT),
    ("profit_factor", "PF", 8, "0.00"),
    ("avg_holding_seconds", "Avg Hold (s)", 12, INT_FMT),
    ("best_trade", "Best", 12, CURRENCY_FMT),
    ("worst_trade", "Worst", 12, CURRENCY_FMT),
    ("running_equity", "Running Equity", 16, CURRENCY_FMT),
]


def _build_daily_summary(ws, daily_df: pd.DataFrame) -> int:
    title_row(ws, "DAILY SUMMARY", len(DAILY_COLS))
    _write_headers(ws, [c[1] for c in DAILY_COLS], 2)

    if daily_df.empty:
        ws.cell(row=3, column=1, value="No data.")
        set_widths(ws, [c[2] for c in DAILY_COLS])
        return 3

    view = daily_df[[c[0] for c in DAILY_COLS if c[0] in daily_df.columns]].copy()
    number_formats = {c[0]: c[3] for c in DAILY_COLS if c[3]}
    last_row = _write_df(ws, view, start_row=3, number_formats=number_formats)

    # Green/red on Net P&L.
    net_idx = [c[0] for c in DAILY_COLS].index("net_profit") + 1
    col_letter = get_column_letter(net_idx)
    ws.conditional_formatting.add(
        f"{col_letter}3:{col_letter}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="63BE7B",
        ),
    )
    end_col_letter = get_column_letter(len(view.columns))
    _add_table(ws, f"A2:{end_col_letter}{last_row}", "DailySummaryTable")

    set_widths(ws, [c[2] for c in DAILY_COLS if c[0] in view.columns])
    ws.freeze_panes = "A3"
    return last_row


def _build_equity_curve(ws, daily_df: pd.DataFrame) -> int:
    title_row(ws, "EQUITY CURVE", 4)
    _write_headers(ws, ["Date", "Net P&L", "Running Equity", "Drawdown"], 2)

    if daily_df.empty:
        ws.cell(row=3, column=1, value="No data.")
        set_widths(ws, [14, 14, 16, 14])
        return 3

    equity = 0.0
    peak = 0.0
    for i, row in daily_df.iterrows():
        r = 3 + i
        date = row["date"]
        net = float(row["net_profit"])
        equity += net
        peak = max(peak, equity)
        dd = equity - peak
        ws.cell(row=r, column=1, value=date).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=net).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3, value=equity).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=dd).number_format = CURRENCY_FMT

    last_row = 2 + len(daily_df)
    apply_band(ws, 3, last_row, 4)

    # Line chart: running equity.
    chart = LineChart()
    chart.title = "Cumulative Net P&L"
    chart.y_axis.title = "Running Equity"
    chart.x_axis.title = "Date"
    chart.height = 10
    chart.width = 22
    data_ref = Reference(ws, min_col=3, min_row=2, max_col=3, max_row=last_row)
    cat_ref = Reference(ws, min_col=1, min_row=3, max_col=1, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, "F3")

    # Drawdown bar below.
    dd_chart = BarChart()
    dd_chart.type = "col"
    dd_chart.title = "Drawdown"
    dd_chart.y_axis.title = "$"
    dd_chart.x_axis.title = "Date"
    dd_chart.height = 7
    dd_chart.width = 22
    dd_ref = Reference(ws, min_col=4, min_row=2, max_col=4, max_row=last_row)
    dd_chart.add_data(dd_ref, titles_from_data=True)
    dd_chart.set_categories(cat_ref)
    ws.add_chart(dd_chart, "F25")

    set_widths(ws, [14, 14, 16, 14])
    ws.freeze_panes = "A3"
    return last_row


def _build_calendar(ws, calendar_df: pd.DataFrame) -> int:
    title_row(ws, "CALENDAR — DAILY NET P&L", max(len(calendar_df.columns) if not calendar_df.empty else 8, 8))
    if calendar_df.empty:
        ws.cell(row=3, column=1, value="No data.")
        set_widths(ws, [14] * 8)
        return 3

    headers = list(calendar_df.columns)
    _write_headers(ws, headers, 2)

    number_formats = {col: CURRENCY_FMT for col in headers if col != "iso_week"}
    last_row = _write_df(ws, calendar_df, start_row=3, number_formats=number_formats)

    # Color-scale the P&L columns (everything except iso_week).
    for c_idx, col in enumerate(headers, start=1):
        if col == "iso_week":
            continue
        col_letter = get_column_letter(c_idx)
        ws.conditional_formatting.add(
            f"{col_letter}3:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFF",
                end_type="max", end_color="63BE7B",
            ),
        )

    set_widths(ws, [14] + [14] * (len(headers) - 1))
    ws.freeze_panes = "B3"
    return last_row


BUCKET_COLS_BASE = [
    ("trades", "Trades", 10, INT_FMT),
    ("wins", "Wins", 8, INT_FMT),
    ("win_rate", "Win %", 10, PERCENT_FMT),
    ("net_profit", "Net P&L", 14, CURRENCY_FMT),
    ("expectancy", "Expectancy", 14, CURRENCY_FMT),
    ("profit_factor", "PF", 8, "0.00"),
    ("avg_holding_seconds", "Avg Hold (s)", 12, INT_FMT),
]


def _build_bucket_sheet(ws, title: str, bucket_col: str, bucket_label: str, bucket_df: pd.DataFrame,
                        chart_title: str | None = None) -> int:
    cols = [(bucket_col, bucket_label, 16, None)] + BUCKET_COLS_BASE
    title_row(ws, title, len(cols))
    _write_headers(ws, [c[1] for c in cols], 2)

    if bucket_df.empty:
        ws.cell(row=3, column=1, value="No data.")
        set_widths(ws, [c[2] for c in cols])
        return 3

    view = bucket_df[[c[0] for c in cols if c[0] in bucket_df.columns]].copy()
    number_formats = {c[0]: c[3] for c in cols if c[3]}
    last_row = _write_df(ws, view, start_row=3, number_formats=number_formats)

    # Net P&L color scale.
    if "net_profit" in view.columns:
        net_idx = list(view.columns).index("net_profit") + 1
        col_letter = get_column_letter(net_idx)
        ws.conditional_formatting.add(
            f"{col_letter}3:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="num", mid_value=0, mid_color="FFFFFF",
                end_type="max", end_color="63BE7B",
            ),
        )

    if chart_title and "net_profit" in view.columns:
        chart = BarChart()
        chart.type = "col"
        chart.title = chart_title
        chart.y_axis.title = "Net P&L"
        chart.x_axis.title = bucket_label
        chart.height = 9
        chart.width = 20
        net_idx = list(view.columns).index("net_profit") + 1
        chart.add_data(Reference(ws, min_col=net_idx, min_row=2, max_col=net_idx, max_row=last_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=3, max_col=1, max_row=last_row))
        anchor_col = get_column_letter(len(view.columns) + 2)
        ws.add_chart(chart, f"{anchor_col}3")

    end_col_letter = get_column_letter(len(view.columns))
    _add_table(ws, f"A2:{end_col_letter}{last_row}", f"{title.replace(' ', '')}Table")
    set_widths(ws, [c[2] for c in cols if c[0] in view.columns])
    ws.freeze_panes = "B3"
    return last_row


def _build_mfe_mae(ws, df: pd.DataFrame) -> int:
    title_row(ws, "MFE / MAE EFFICIENCY", 5)
    _write_headers(ws, ["Symbol", "Side", "Price Move", "MFE", "MAE"], 2)

    if df.empty:
        ws.cell(row=3, column=1, value="No data.")
        set_widths(ws, [12, 8, 14, 14, 14])
        return 3

    view = df[["symbol", "direction_str", "price_move", "mfe", "mae"]].copy()
    number_formats = {"price_move": PRICE_FMT, "mfe": PRICE_FMT, "mae": PRICE_FMT}
    last_row = _write_df(ws, view, start_row=3, number_formats=number_formats)

    # Summary stats bottom.
    summary_row = last_row + 2
    ws.cell(row=summary_row, column=1, value="Avg MFE").font = BOLD_FONT
    ws.cell(row=summary_row, column=2, value=float(df["mfe"].mean(skipna=True)) if df["mfe"].notna().any() else 0).number_format = PRICE_FMT
    ws.cell(row=summary_row + 1, column=1, value="Avg MAE").font = BOLD_FONT
    ws.cell(row=summary_row + 1, column=2, value=float(df["mae"].mean(skipna=True)) if df["mae"].notna().any() else 0).number_format = PRICE_FMT
    ws.cell(row=summary_row + 2, column=1, value="Avg MFE Efficiency").font = BOLD_FONT
    ws.cell(row=summary_row + 2, column=2, value=float(df["mfe_efficiency"].mean(skipna=True)) if df["mfe_efficiency"].notna().any() else 0).number_format = PERCENT_FMT

    # Scatter: MFE vs price move.
    chart = ScatterChart()
    chart.title = "Realised Move vs MFE"
    chart.style = 13
    chart.x_axis.title = "MFE"
    chart.y_axis.title = "Realised Move"
    chart.height = 10
    chart.width = 18
    x_ref = Reference(ws, min_col=4, min_row=3, max_row=last_row)
    y_ref = Reference(ws, min_col=3, min_row=3, max_row=last_row)
    series = Series(y_ref, x_ref, title="Trades")
    chart.series.append(series)
    ws.add_chart(chart, "G3")

    set_widths(ws, [12, 8, 14, 14, 14])
    ws.freeze_panes = "A3"
    return summary_row + 2


def _build_risk_metrics(ws, stats: transform.JournalStats, currency: str) -> int:
    title_row(ws, "RISK METRICS", 2)
    metrics = [
        ("Total Trades", stats.trades, INT_FMT),
        ("Wins", stats.wins, INT_FMT),
        ("Losses", stats.losses, INT_FMT),
        ("Win Rate", stats.win_rate, PERCENT_FMT),
        ("Gross Winning P&L", stats.gross_profit, CURRENCY_FMT),
        ("Gross Losing P&L", stats.gross_loss, CURRENCY_FMT),
        ("Net P&L", stats.net_profit, CURRENCY_FMT),
        ("Profit Factor", stats.profit_factor if stats.profit_factor != float("inf") else "∞", "0.00"),
        ("Expectancy / Trade", stats.expectancy, CURRENCY_FMT),
        ("Avg Win", stats.avg_win, CURRENCY_FMT),
        ("Avg Loss", stats.avg_loss, CURRENCY_FMT),
        ("Best Trade", stats.best_trade, CURRENCY_FMT),
        ("Worst Trade", stats.worst_trade, CURRENCY_FMT),
        ("Max Drawdown", stats.max_drawdown, CURRENCY_FMT),
        ("Max Consecutive Losses", stats.max_consec_losses, INT_FMT),
        ("Sharpe of Trades", stats.sharpe_of_trades, "0.00"),
        ("SL-Hit %", stats.sl_hit_pct, PERCENT_FMT),
        ("TP-Hit %", stats.tp_hit_pct, PERCENT_FMT),
        ("Commission / Swap Drag %", stats.commission_drag_pct, PERCENT_FMT),
        ("Account Currency", currency, None),
    ]
    for i, (label, val, fmt) in enumerate(metrics, start=3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.fill = SUBHEADER_FILL
        label_cell.font = SUBHEADER_FONT
        label_cell.border = BORDER
        label_cell.alignment = LEFT
        val_cell = ws.cell(row=i, column=2, value=val)
        val_cell.fill = PatternFill("solid", start_color=REY_BLUE_LIGHT)
        val_cell.font = BASE_FONT
        val_cell.alignment = CENTER
        val_cell.border = BORDER
        if fmt:
            val_cell.number_format = fmt

    set_widths(ws, [30, 24])
    return len(metrics) + 2


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_workbook(df: pd.DataFrame, out_path: Path, generated_at: datetime | None = None,
                   source_label: str = "MT5") -> Path:
    """Render the journal workbook from a normalised DataFrame."""
    generated_at = generated_at or datetime.utcnow()

    daily_df = transform.daily_summary(df)
    cal_df = transform.calendar_pivot(df)
    hour_df = transform.by_hour(df)
    session_df = transform.by_session(df)
    weekday_df = transform.by_weekday(df)
    symbol_df = transform.by_symbol(df)
    account_df = transform.by_account(df)
    holding_df = transform.by_holding_bucket(df)
    stats = transform.JournalStats.compute(df)

    # Cover metadata.
    if not df.empty:
        account_list = sorted(df["account_id"].astype(str).unique().tolist())
        start_dt = df["exit_time"].min()
        end_dt = df["exit_time"].max()
        date_range = f"{start_dt:%Y-%m-%d} → {end_dt:%Y-%m-%d}"
    else:
        account_list = []
        date_range = "—"

    meta = {
        "generated_at": generated_at.strftime("%Y-%m-%d %H:%M:%S"),
        "source": source_label,
        "accounts": ", ".join(account_list) if account_list else "—",
        "date_range": date_range,
        "total_trades": int(stats.trades),
        "net_profit_str": f"{stats.net_profit:,.2f}",
        "currency": "USD",
    }

    wb = Workbook()
    wb.remove(wb.active)

    sheets: list[tuple[str, int, callable]] = [
        ("Cover", 4, lambda ws: _build_cover(ws, meta)),
        ("About", 2, lambda ws: _build_about(ws)),
        ("Trade Log", len(TRADE_LOG_COLS), lambda ws: _build_trade_log(ws, df)),
        ("Daily Summary", len(DAILY_COLS), lambda ws: _build_daily_summary(ws, daily_df)),
        ("Equity Curve", 4, lambda ws: _build_equity_curve(ws, daily_df)),
        ("Calendar", max(len(cal_df.columns) if not cal_df.empty else 8, 8),
         lambda ws: _build_calendar(ws, cal_df)),
        ("By Hour", 1 + len(BUCKET_COLS_BASE),
         lambda ws: _build_bucket_sheet(ws, "BY HOUR OF DAY (UTC)", "hour_of_day", "Hour", hour_df,
                                       chart_title="Net P&L by Hour")),
        ("By Session", 1 + len(BUCKET_COLS_BASE),
         lambda ws: _build_bucket_sheet(ws, "BY SESSION", "session", "Session", session_df,
                                       chart_title="Net P&L by Session")),
        ("By Symbol", 1 + len(BUCKET_COLS_BASE),
         lambda ws: _build_bucket_sheet(ws, "BY SYMBOL", "symbol", "Symbol", symbol_df,
                                       chart_title="Net P&L by Symbol")),
        ("By Weekday", 1 + len(BUCKET_COLS_BASE),
         lambda ws: _build_bucket_sheet(ws, "BY WEEKDAY", "weekday", "Weekday", weekday_df,
                                       chart_title="Net P&L by Weekday")),
        ("Holding Buckets", 1 + len(BUCKET_COLS_BASE),
         lambda ws: _build_bucket_sheet(ws, "HOLDING TIME BUCKETS", "holding_bucket", "Bucket", holding_df,
                                       chart_title="Net P&L by Holding Bucket")),
        ("MFE-MAE", 5, lambda ws: _build_mfe_mae(ws, df)),
        ("Risk Metrics", 2, lambda ws: _build_risk_metrics(ws, stats, meta["currency"])),
        ("By Account", 1 + len(BUCKET_COLS_BASE),
         lambda ws: _build_bucket_sheet(ws, "BY ACCOUNT", "account_id", "Account", account_df,
                                       chart_title="Net P&L by Account")),
    ]

    for name, ncols, builder in sheets:
        ws = wb.create_sheet(name)
        builder(ws)
        add_confidential_strip(ws, ncols)
        add_logo(ws, anchor="A1")
        add_confidential_header(ws)
        if ws.freeze_panes:
            fp = ws.freeze_panes
            col_letters = "".join(ch for ch in fp if ch.isalpha())
            row_num = int("".join(ch for ch in fp if ch.isdigit()))
            ws.freeze_panes = f"{col_letters}{row_num + 3}"

    wb.properties.title = "Rey Capital — Daily Scalping Trading Journal"
    wb.properties.creator = "Rey Capital"
    wb.properties.company = "Rey Capital"
    wb.properties.keywords = "CONFIDENTIAL, Scalping, Trading Journal, MT5, Rey Capital"
    wb.properties.description = "Confidential — auto-generated scalping trading journal."

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
