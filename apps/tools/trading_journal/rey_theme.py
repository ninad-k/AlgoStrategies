"""Rey Capital palette, fonts, and shared openpyxl helpers.

Mirrors the constants in docs/strategy-intake/_build_xlsx.py so the journal
keeps the same visual identity. Kept local (rather than cross-importing from
docs/strategy-intake/, which has a hyphen in the path and is not a Python
package) to avoid sys.path gymnastics.
"""
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FONT_NAME = "Calibri"

REY_BLUE = "1E3FAE"
REY_BLUE_DARK = "10267A"
REY_BLUE_MID = "3457C7"
REY_BLUE_LIGHT = "E4EAFB"
CONFIDENTIAL_RED = "C00000"
POSITIVE_GREEN = "2E7D32"
NEGATIVE_RED = "C62828"
NEUTRAL_AMBER = "F9A825"

HEADER_FILL = PatternFill("solid", start_color=REY_BLUE)
SUBHEADER_FILL = PatternFill("solid", start_color=REY_BLUE_MID)
BAND_FILL = PatternFill("solid", start_color=REY_BLUE_LIGHT)
TITLE_FILL = PatternFill("solid", start_color=REY_BLUE_DARK)
CONFIDENTIAL_FILL = PatternFill("solid", start_color="FDE7E9")

HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)
CONFIDENTIAL_FONT = Font(name=FONT_NAME, bold=True, color=CONFIDENTIAL_RED, size=10, italic=True)
BASE_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)

THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

REPO_ROOT = Path(__file__).resolve().parents[3]
LOGO_PATH = REPO_ROOT / "docs" / "strategy-intake" / "ReyCapital_Logo.png"


def title_row(ws, text: str, ncols: int, row: int = 1) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 34


def style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def apply_band(ws, start_row: int, end_row: int, ncols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else CENTER
            if (r - start_row) % 2 == 1:
                cell.fill = BAND_FILL


def set_widths(ws, widths) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_confidential_header(ws, footer_right: str = "Rey Capital Scalping Journal v1") -> None:
    ws.oddHeader.center.text = "CONFIDENTIAL — REY CAPITAL INTERNAL USE ONLY"
    ws.oddHeader.center.size = 11
    ws.oddHeader.center.color = CONFIDENTIAL_RED
    ws.oddHeader.left.text = "Rey Capital | Smart Investments"
    ws.oddHeader.left.size = 9
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.center.size = 9
    ws.oddFooter.right.text = footer_right
    ws.oddFooter.right.size = 9


def add_logo(ws, anchor: str = "A1", width: int = 180, height: int = 55) -> bool:
    """Insert the Rey Capital logo if the asset + Pillow are available."""
    if not LOGO_PATH.exists():
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage

        img = XLImage(str(LOGO_PATH))
    except ImportError:
        # Pillow not installed — skip the logo rather than fail the whole build.
        return False
    img.width = width
    img.height = height
    img.anchor = anchor
    ws.add_image(img)
    return True


def add_confidential_strip(ws, ncols: int) -> None:
    """Insert a 3-row CONFIDENTIAL banner strip at the top of the sheet."""
    ws.insert_rows(1, amount=3)
    ws.row_dimensions[1].height = 50
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ncols, 4))
    cell = ws.cell(row=2, column=1, value="CONFIDENTIAL — REY CAPITAL | SMART INVESTMENTS")
    cell.fill = CONFIDENTIAL_FILL
    cell.font = CONFIDENTIAL_FONT
    cell.alignment = CENTER
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6
