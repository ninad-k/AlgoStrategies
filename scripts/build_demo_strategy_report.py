"""Demo-account variant of build_strategy_report.

Differences vs the live builder:
- Reads from C:/Users/Ninad/Downloads/Demo Reports
- Strategy detection uses the FULL order-comment text, not just (parens) tags.
  Buy/Sell/BUY/SELL suffixes are stripped so a single strategy is not split.
- Any trade whose comment is missing, numeric, an SL/TP echo, or a {magic}
  token is labelled "William Larry" (i.e. a 'magic number only' trade).
"""
from __future__ import annotations
import os, re, sys
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

# Reuse styling/aggregation/writing helpers from the live builder.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_strategy_report as base

sys.stdout.reconfigure(encoding="utf-8")

REPORTS_DIR = r"C:\Users\Ninad\Downloads\Demo Reports"
OUT_PATH    = r"C:\Users\Ninad\OneDrive - Reycapitals\Desktop\demo_strategy_analysis.xlsx"

DIRECTION_RE = re.compile(r"\s*(buy|sell|long|short)\s*$", re.IGNORECASE)
NUMERIC_RE   = re.compile(r"^[\s\{\[]*[\d\.\-]+[\s\}\]]*$")
SLTP_RE      = re.compile(r"^\[(sl|tp)\b", re.IGNORECASE)

def classify(comment: str | None) -> str:
    if not comment:
        return "William Larry"
    c = str(comment).strip()
    if not c:
        return "William Larry"
    if SLTP_RE.match(c):
        return "William Larry"
    if NUMERIC_RE.match(c):
        return "William Larry"
    if c.startswith("{") and c.rstrip().endswith("}"):
        return "William Larry"
    # Strip trailing direction word so "Renko V3 Buy"/"Renko V3 Sell" merge.
    cleaned = DIRECTION_RE.sub("", c).strip()
    cleaned = re.sub(r"\s+(ENTRY|Entry|Order|ORDER)$", "", cleaned).strip()
    return cleaned or "William Larry"

def parse_report(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    section_at = {}
    for i, row in enumerate(rows, start=1):
        v = row[0] if row else None
        if v in ("Positions", "Orders", "Deals", "Results"):
            section_at[v] = i

    pos_start = section_at.get("Positions")
    ord_start = section_at.get("Orders")
    if not pos_start or not ord_start:
        return []

    # order_id -> comment
    orders_comment = {}
    end = section_at.get("Deals", section_at.get("Results", len(rows) + 1))
    for i in range(ord_start + 1, end - 1):
        row = rows[i]
        if not row or not row[0]:
            continue
        oid = row[1]
        cm = row[11] if len(row) > 11 else None
        if oid is not None and cm:
            orders_comment[oid] = str(cm)

    trades = []
    pos_end = ord_start
    for i in range(pos_start + 1, pos_end - 1):
        row = rows[i]
        if not row or row[0] in (None, "", "Positions", "Orders"):
            continue
        open_time = base.parse_dt(row[0])
        if not open_time:
            continue
        position_id = row[1]
        symbol      = row[2] or ""
        ttype       = row[3] or ""
        volume      = base.to_float(row[4])
        close_time  = base.parse_dt(row[8])
        commission  = base.to_float(row[10])
        swap        = base.to_float(row[11])
        profit      = base.to_float(row[12])
        net = profit + commission + swap

        comment  = orders_comment.get(position_id, "")
        strategy = classify(comment)

        trades.append({
            "open_time": open_time,
            "close_time": close_time or open_time,
            "date": (close_time or open_time).date(),
            "symbol": symbol,
            "type": ttype,
            "volume": volume,
            "profit": net,
            "strategy": strategy,
        })
    return trades

def load_all():
    out = []
    for f in sorted(os.listdir(REPORTS_DIR)):
        if not f.lower().endswith(".xlsx"): continue
        p = os.path.join(REPORTS_DIR, f)
        print(f"Reading {f} ...", flush=True)
        t = parse_report(p)
        print(f"  -> {len(t)} positions", flush=True)
        out.extend(t)
    return out

def main():
    trades = load_all()
    print(f"Total positions: {len(trades)}")
    if not trades:
        print("No trades."); return

    daily     = base.daily_summary(trades)
    strat_agg = base.strategy_aggregate(trades)
    total_net = sum(t["profit"] for t in trades)
    dates     = sorted({t["date"] for t in trades})
    strats    = sorted({t["strategy"] for t in trades})
    print(f"Strategies ({len(strats)}): {strats}")
    print(f"Date range: {dates[0]} -> {dates[-1]}")

    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    base.write_daily_summary(wb, daily, total_net, dates, strats)
    base.write_strategy_analysis(wb, strat_agg, dates[0], dates[-1])
    wb.save(OUT_PATH)
    print(f"Saved -> {OUT_PATH}")

if __name__ == "__main__":
    main()
