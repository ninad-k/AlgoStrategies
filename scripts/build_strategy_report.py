"""Build a strategy_analysis_*.xlsx report from MT5 trade history exports.

Mirrors the template at strategy_analysis_4.xlsx (Daily Summary + Strategy Analysis sheets).
Strategy tag is parsed from the entry order's comment, e.g. "(combo)" -> "combo".
"""
from __future__ import annotations
import os, re, sys, io
from datetime import datetime, date
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

sys.stdout.reconfigure(encoding="utf-8")

REPORTS_DIR = r"C:\Users\Ninad\Downloads\Reports"
OUT_PATH    = r"C:\Users\Ninad\OneDrive - Reycapitals\Desktop\strategy_analysis_5.xlsx"
# Optional whitelist (comma-separated env var) to filter strategies; empty = all.
STRATEGY_FILTER = set(filter(None, os.environ.get("STRAT_FILTER", "").split(",")))

STRAT_RE = re.compile(r"\(([A-Za-z0-9_]+)")  # first parenthesised tag

def parse_dt(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        for fmt in ("%Y.%m.%d %H:%M:%S", "%Y.%m.%d %H:%M"):
            try: return datetime.strptime(v, fmt)
            except ValueError: pass
    return None

def to_float(v):
    if v is None or v == "": return 0.0
    try: return float(v)
    except (TypeError, ValueError):
        try: return float(str(v).replace(" ", "").replace(",", ""))
        except ValueError: return 0.0

def find_section(ws, name):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v == name:
            return r
    return None

def parse_report(path):
    """Return list of trade dicts from a single MT5 export."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Locate sections
    section_at = {}
    for i, row in enumerate(rows, start=1):
        v = row[0] if row else None
        if v in ("Positions", "Orders", "Deals", "Results"):
            section_at[v] = i

    pos_start = section_at.get("Positions")
    ord_start = section_at.get("Orders")
    if not pos_start or not ord_start:
        return []

    # Build order_id -> comment map (use Orders' Comment column = idx 11, 0-based)
    orders_comment = {}
    end = section_at.get("Deals", section_at.get("Results", len(rows) + 1))
    for i in range(ord_start + 1, end - 1):
        row = rows[i]  # 0-based index
        if not row or not row[0]:
            continue
        oid = row[1]
        if oid is None: continue
        cm = row[11] if len(row) > 11 else None
        if cm:
            orders_comment[oid] = str(cm)

    # Iterate Positions section trades
    trades = []
    pos_end = ord_start
    for i in range(pos_start + 1, pos_end - 1):
        row = rows[i]  # 0-based
        if not row or row[0] in (None, "", "Positions", "Orders"):
            continue
        open_time = parse_dt(row[0])
        if not open_time:
            continue
        position_id = row[1]
        symbol      = row[2]
        ttype       = row[3]
        volume      = to_float(row[4])
        close_time  = parse_dt(row[8])
        commission  = to_float(row[10])
        swap        = to_float(row[11])
        profit      = to_float(row[12])
        net = profit + commission + swap

        comment = orders_comment.get(position_id, "")
        m = STRAT_RE.search(comment)
        if m:
            strategy = m.group(1)
        elif comment.startswith("QQ["):
            strategy = "QQGrid"
        elif "Flask" in comment:
            strategy = "Flask"
        elif comment.startswith("EMA Cross"):
            strategy = "EMACross"
        else:
            strategy = "Other"

        trades.append({
            "open_time": open_time,
            "close_time": close_time or open_time,
            "date": (close_time or open_time).date(),
            "symbol": symbol or "",
            "type": ttype or "",
            "volume": volume,
            "profit": net,
            "strategy": strategy,
        })
    return trades

def load_all_trades():
    trades = []
    for f in sorted(os.listdir(REPORTS_DIR)):
        if not f.lower().endswith(".xlsx"): continue
        path = os.path.join(REPORTS_DIR, f)
        print(f"Reading {f} ...", flush=True)
        t = parse_report(path)
        print(f"  -> {len(t)} positions", flush=True)
        trades.extend(t)
    return trades

# ---------- Aggregations ----------

def daily_summary(trades):
    """key=(date, strategy, symbol) -> aggregates"""
    agg = defaultdict(lambda: {"trades": 0, "wins": 0, "losses": 0,
                                "gp": 0.0, "gl": 0.0, "best": None, "worst": None})
    for t in trades:
        k = (t["date"], t["strategy"], t["symbol"])
        a = agg[k]
        a["trades"] += 1
        p = t["profit"]
        if p > 0:
            a["wins"] += 1; a["gp"] += p
        elif p < 0:
            a["losses"] += 1; a["gl"] += p
        a["best"]  = p if a["best"]  is None or p > a["best"]  else a["best"]
        a["worst"] = p if a["worst"] is None or p < a["worst"] else a["worst"]
    return agg

def strategy_aggregate(trades):
    agg = defaultdict(lambda: {
        "start": None, "instruments": set(), "lots": set(),
        "trades": 0, "wins": 0, "losses": 0, "gp": 0.0, "gl": 0.0,
        "win_amts": [], "loss_amts": [],
        "best": None, "worst": None,
        "long_trades": 0, "long_wins": 0, "long_pnl": 0.0,
        "short_trades": 0, "short_wins": 0, "short_pnl": 0.0,
        "monthly": defaultdict(float),
        "running_pnls": [],
    })
    by_strat = defaultdict(list)
    for t in trades:
        by_strat[t["strategy"]].append(t)
    for s, lst in by_strat.items():
        lst.sort(key=lambda x: x["close_time"])
        a = agg[s]
        running = 0.0; peak = 0.0; max_dd = 0.0
        for t in lst:
            if a["start"] is None or t["open_time"] < a["start"]:
                a["start"] = t["open_time"]
            a["instruments"].add(t["symbol"])
            a["lots"].add(round(t["volume"], 2))
            a["trades"] += 1
            p = t["profit"]
            if p > 0:
                a["wins"] += 1; a["gp"] += p; a["win_amts"].append(p)
            elif p < 0:
                a["losses"] += 1; a["gl"] += p; a["loss_amts"].append(p)
            a["best"]  = p if a["best"]  is None or p > a["best"]  else a["best"]
            a["worst"] = p if a["worst"] is None or p < a["worst"] else a["worst"]
            if t["type"] == "buy":
                a["long_trades"] += 1
                if p > 0: a["long_wins"] += 1
                a["long_pnl"] += p
            elif t["type"] == "sell":
                a["short_trades"] += 1
                if p > 0: a["short_wins"] += 1
                a["short_pnl"] += p
            ym = t["close_time"].strftime("%Y-%m")
            a["monthly"][ym] += p
            running += p
            peak = max(peak, running)
            dd = running - peak
            if dd < max_dd: max_dd = dd
        a["max_dd"] = max_dd
    return agg

# ---------- Excel writing ----------

ARIAL_BOLD = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ARIAL = Font(name="Arial", size=10)
HDR_FILL = PatternFill("solid", start_color="305496")
SUB_FILL = PatternFill("solid", start_color="D9E1F2")
TOTAL_FILL = PatternFill("solid", start_color="FFE699")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_2 = '#,##0.00;(#,##0.00);"-"'
NUM_PCT = '0.0"%"'

def style_header(cell):
    cell.font = ARIAL_BOLD
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = BORDER

def write_daily_summary(wb, daily, total_net, dates, strategies):
    ws = wb.create_sheet("Daily Summary")
    if not dates:
        ws["A1"] = "No data"
        return
    dmin, dmax = min(dates), max(dates)
    ws["A1"] = f"Daily P&L by Strategy — {dmin.strftime('%b %d')} to {dmax.strftime('%b %d, %Y')}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws.merge_cells("A1:L1")

    ws["A2"] = "Total Net PnL:"; ws["A2"].font = Font(name="Arial", bold=True)
    ws["C2"] = total_net; ws["C2"].number_format = NUM_2; ws["C2"].font = Font(name="Arial", bold=True, color="0070C0")
    ws["E2"] = f"Days: {len(dates)}"
    ws["G2"] = f"Strategies: {len(strategies)}"
    for c in ("E2","G2"): ws[c].font = Font(name="Arial", bold=True)

    headers = ["Date","Strategy","Instrument","Trades","Wins","Losses","Win Rate",
               "Gross Profit","Gross Loss","Net PnL","Best Trade","Worst Trade"]
    for i,h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h); style_header(c)

    keys = sorted(daily.keys(), key=lambda k: (k[0], k[1], k[2]))
    r = 5
    prev_date = None
    for k in keys:
        d, strat, sym = k
        a = daily[k]
        is_new = d != prev_date
        date_val = d if is_new else None
        prev_date = d
        gp = a["gp"]; gl = a["gl"]; net = gp + gl
        wr = (a["wins"] / a["trades"] * 100) if a["trades"] else 0
        row_vals = [date_val, strat, sym, a["trades"], a["wins"], a["losses"], wr,
                    gp, gl, net, a["best"] or 0, a["worst"] or 0]
        for i,v in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = ARIAL; c.border = BORDER
            if i == 1 and v is not None:
                c.number_format = "yyyy-mm-dd"; c.font = Font(name="Arial", bold=True)
            elif i == 7:
                c.number_format = NUM_PCT
            elif i in (8,9,10,11,12):
                c.number_format = NUM_2
                if i == 10:
                    c.font = Font(name="Arial", bold=True, color="00B050" if net>=0 else "C00000")
            elif i in (4,5,6):
                c.alignment = CENTER
        r += 1

    widths = [13, 18, 14, 9, 8, 9, 11, 13, 13, 13, 13, 13]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

def write_strategy_analysis(wb, strat_agg, dmin, dmax):
    ws = wb.create_sheet("Strategy Analysis")
    ws["A1"] = f"Strategy Performance — {dmin.strftime('%b %d')} to {dmax.strftime('%b %d, %Y')}"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws.merge_cells("A1:T1")

    headers = ["Strategy","Start Date","Lot Size","Instrument","Trades","Wins","Losses","Win Rate",
               "Gross Profit","Gross Loss","Net PnL","Avg Win","Avg Loss","Profit Factor",
               "Expectancy","Max DD","Best Trade","Worst Trade","YTD %","CAGR %"]
    for i,h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h); style_header(c)

    items = sorted(strat_agg.items(), key=lambda kv: -kv[1]["gp"] - kv[1]["gl"])
    r = 4
    tot_trades = tot_wins = tot_losses = 0
    tot_gp = tot_gl = 0.0
    BASE_CAPITAL = 1000.0  # baseline used for YTD% — informational only
    for s, a in items:
        net = a["gp"] + a["gl"]
        wr = (a["wins"]/a["trades"]*100) if a["trades"] else 0
        avg_w = (sum(a["win_amts"])/len(a["win_amts"])) if a["win_amts"] else 0
        avg_l = (sum(a["loss_amts"])/len(a["loss_amts"])) if a["loss_amts"] else 0
        pf = (a["gp"] / -a["gl"]) if a["gl"] else (float("inf") if a["gp"] else 0)
        if pf == float("inf"): pf_v = "∞"
        else: pf_v = round(pf, 2)
        exp = (net / a["trades"]) if a["trades"] else 0
        lots = sorted(a["lots"])
        lot_s = "/".join(str(x) for x in lots) if lots else ""
        instr = ", ".join(sorted(x for x in a["instruments"] if x))
        ytd = net / BASE_CAPITAL * 100
        # naive CAGR over period
        if a["start"] and dmax:
            days = max((dmax - a["start"].date()).days, 1)
            try:
                cagr = ((1 + net/BASE_CAPITAL) ** (365/days) - 1) * 100 if (1 + net/BASE_CAPITAL) > 0 else None
            except (OverflowError, ValueError):
                cagr = None
        else:
            cagr = None

        row_vals = [s, a["start"], lot_s, instr, a["trades"], a["wins"], a["losses"], wr,
                    a["gp"], a["gl"], net, avg_w, avg_l, pf_v, exp, a["max_dd"],
                    a["best"] or 0, a["worst"] or 0, ytd, cagr]
        for i,v in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=i, value=v); c.font = ARIAL; c.border = BORDER
            if i == 2 and isinstance(v, datetime):
                c.number_format = "yyyy-mm-dd"
            elif i == 8 or i in (19, 20):
                c.number_format = NUM_PCT
            elif i in (5,6,7):
                c.alignment = CENTER
            elif i in (9,10,11,12,13,15,16,17,18):
                c.number_format = NUM_2
                if i == 11:
                    c.font = Font(name="Arial", bold=True, color="00B050" if net>=0 else "C00000")
        tot_trades += a["trades"]; tot_wins += a["wins"]; tot_losses += a["losses"]
        tot_gp += a["gp"]; tot_gl += a["gl"]
        r += 1

    # TOTAL row using formulas
    last_data_row = r - 1
    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
    for col in (5,6,7,9,10,11):
        col_l = get_column_letter(col)
        f = f"=SUM({col_l}4:{col_l}{last_data_row})"
        c = ws.cell(row=total_row, column=col, value=f)
        c.font = Font(name="Arial", bold=True)
        c.fill = TOTAL_FILL
        if col >= 9:
            c.number_format = NUM_2
        else:
            c.alignment = CENTER
    # Win rate total
    ws.cell(row=total_row, column=8, value=f"=IFERROR(F{total_row}/E{total_row}*100,0)").number_format = NUM_PCT
    ws.cell(row=total_row, column=8).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    for col in range(1, 21):
        ws.cell(row=total_row, column=col).border = BORDER

    # ----- Monthly Breakdown -----
    r += 2
    ws.cell(row=r, column=1, value="Monthly Breakdown by Strategy").font = Font(name="Arial", size=12, bold=True)
    r += 1
    months = sorted({m for a in strat_agg.values() for m in a["monthly"].keys()})
    headers2 = ["Strategy"] + [""]*2 + months + ["Total"]
    for i,h in enumerate(headers2, start=1):
        c = ws.cell(row=r, column=i, value=h)
        if h: style_header(c)
    r += 1
    for s, a in items:
        ws.cell(row=r, column=1, value=s).font = ARIAL
        for i,m in enumerate(months):
            v = a["monthly"].get(m, 0)
            c = ws.cell(row=r, column=4+i, value=v)
            c.font = ARIAL; c.number_format = NUM_2; c.border = BORDER
        # Total formula
        if months:
            first = get_column_letter(4)
            last = get_column_letter(4+len(months)-1)
            c = ws.cell(row=r, column=4+len(months), value=f"=SUM({first}{r}:{last}{r})")
            c.font = Font(name="Arial", bold=True); c.number_format = NUM_2; c.border = BORDER
        ws.cell(row=r, column=1).border = BORDER
        r += 1

    # ----- Long vs Short -----
    r += 2
    ws.cell(row=r, column=1, value="Long vs Short by Strategy").font = Font(name="Arial", size=12, bold=True)
    r += 1
    headers3 = ["Strategy","","","Long Trades","Long WR","Long PnL","Short Trades","Short WR","Short PnL"]
    for i,h in enumerate(headers3, start=1):
        c = ws.cell(row=r, column=i, value=h)
        if h: style_header(c)
    r += 1
    for s, a in sorted(strat_agg.items()):
        long_wr = (a["long_wins"]/a["long_trades"]*100) if a["long_trades"] else 0
        short_wr = (a["short_wins"]/a["short_trades"]*100) if a["short_trades"] else 0
        row_vals = [s, None, None, a["long_trades"], long_wr, a["long_pnl"],
                    a["short_trades"], short_wr, a["short_pnl"]]
        for i,v in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=i, value=v); c.font = ARIAL; c.border = BORDER
            if i in (5,8): c.number_format = NUM_PCT
            elif i in (6,9): c.number_format = NUM_2
            elif i in (4,7): c.alignment = CENTER
        r += 1

    widths = [16,12,11,14,8,9,9,11,13,13,13,11,11,13,12,13,13,13,11,12]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

def main():
    trades = load_all_trades()
    print(f"Total positions parsed: {len(trades)}")
    if STRATEGY_FILTER:
        trades = [t for t in trades if t["strategy"] in STRATEGY_FILTER]
        print(f"Filtered to {STRATEGY_FILTER}: {len(trades)} positions")
    if not trades:
        print("No trades found, exiting."); return
    daily = daily_summary(trades)
    strat_agg = strategy_aggregate(trades)
    total_net = sum(t["profit"] for t in trades)
    dates = sorted({t["date"] for t in trades})
    strategies = sorted({t["strategy"] for t in trades})
    print(f"Strategies: {strategies}")
    print(f"Date range: {dates[0]} to {dates[-1]}")

    wb = Workbook()
    wb.remove(wb.active)
    write_daily_summary(wb, daily, total_net, dates, strategies)
    write_strategy_analysis(wb, strat_agg, dates[0], dates[-1])
    wb.save(OUT_PATH)
    print(f"Saved -> {OUT_PATH}")

if __name__ == "__main__":
    main()
