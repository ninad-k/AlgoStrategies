"""Build CFD Strategy Params workbook - blank template + filled sample. Rey Capital themed."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.header_footer import HeaderFooter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from copy import copy

OUT_DIR = r"C:/Users/NinadKulkarni/PycharmProjects/AlgoStrategies/docs/strategy-intake"
LOGO_PATH = f"{OUT_DIR}/ReyCapital_Logo.png"

# ---- Rey Capital theme ----
FONT_NAME = "Calibri"
REY_BLUE = "1E3FAE"        # primary brand blue
REY_BLUE_DARK = "10267A"   # deep navy for banners
REY_BLUE_MID = "3457C7"
REY_BLUE_LIGHT = "E4EAFB"  # band fill
CONFIDENTIAL_RED = "C00000"

HEADER_FILL = PatternFill("solid", start_color=REY_BLUE)
SUBHEADER_FILL = PatternFill("solid", start_color=REY_BLUE_MID)
BAND_FILL = PatternFill("solid", start_color=REY_BLUE_LIGHT)
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")
TITLE_FILL = PatternFill("solid", start_color=REY_BLUE_DARK)
CONFIDENTIAL_FILL = PatternFill("solid", start_color="FDE7E9")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)
CONFIDENTIAL_FONT = Font(name=FONT_NAME, bold=True, color=CONFIDENTIAL_RED, size=10, italic=True)
BASE_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_logo(ws, anchor="A1"):
    """Embed Rey Capital logo anchored at given cell."""
    img = XLImage(LOGO_PATH)
    img.width = 180
    img.height = 55
    img.anchor = anchor
    ws.add_image(img)


def add_confidential_header(ws):
    """Set page header/footer with CONFIDENTIAL marking."""
    ws.oddHeader.center.text = "CONFIDENTIAL — REY CAPITAL INTERNAL USE ONLY"
    ws.oddHeader.center.size = 11
    ws.oddHeader.center.color = CONFIDENTIAL_RED
    ws.oddHeader.left.text = "Rey Capital | Smart Investments"
    ws.oddHeader.left.size = 9
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.center.size = 9
    ws.oddFooter.right.text = "CFD Strategy Intake v1.0"
    ws.oddFooter.right.size = 9


# ---- Dropdown vocabularies ----
DV_ASSET_CLASS = '"FX Major,FX Minor,FX Exotic,Index CFD,Commodity CFD,Crypto CFD,Stock CFD,Bond CFD"'
DV_SESSION = '"Sydney,Tokyo,London,New York,London-NY Overlap,Tokyo-London Overlap,24/5,24/7"'
DV_ORDER_TYPE = '"Market,Limit,Stop,Stop-Limit,SL-M,Iceberg,TWAP,VWAP"'
DV_YESNO = '"Yes,No"'
DV_DIRECTION = '"Long Only,Short Only,Both"'
DV_TIMEFRAME = '"Tick,M1,M5,M15,M30,H1,H4,D1,W1"'
DV_CURRENCY = '"USD,EUR,GBP,JPY,INR,AUD,CAD,CHF"'


def style_header(ws, row, ncols, text_map=None, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        if text_map and c in text_map:
            cell.value = text_map[c]
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.fill = TITLE_FILL
    c.font = TITLE_FONT
    c.alignment = CENTER
    ws.row_dimensions[1].height = 34


def apply_band(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = LEFT if c == 1 else CENTER
            if (r - start_row) % 2 == 1:
                cell.fill = BAND_FILL


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, formula, cell_range):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.error = "Please pick a value from the dropdown"
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(cell_range)


# ==========================================================================
# SHEET BUILDERS
# ==========================================================================

def build_cover(ws, sample=False):
    title_row(ws, "CFD STRATEGY PARAMETERS — DHRUVA Automation Intake", 4)
    rows = [
        ("Strategy Name",        "EMA 9-21 Crossover (FX & Indices)" if sample else ""),
        ("Strategy Code",        "CFD_EMA_9_21_v1" if sample else ""),
        ("Version",              "1.0" if sample else ""),
        ("Author / PM",          "Ninad Kulkarni" if sample else ""),
        ("Desk",                 "Systematic FX Desk" if sample else ""),
        ("Submission Date",      "2026-04-20" if sample else ""),
        ("Base Currency",        "USD" if sample else ""),
        ("Direction",            "Both" if sample else ""),
        ("Primary Timeframe",    "H1" if sample else ""),
        ("Strategy Category",    "Trend-Following" if sample else ""),
        ("Companion Word Doc",   "CFD_Strategy_Template_SAMPLE_EMA_9_21.docx" if sample else ""),
        ("Status",               "DRAFT — Pending Sign-off" if sample else ""),
    ]
    for i, (label, val) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws.cell(row=i, column=1).fill = SUBHEADER_FILL
        ws.cell(row=i, column=1).font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
        ws.cell(row=i, column=1).alignment = LEFT
        ws.cell(row=i, column=1).border = BORDER
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        v = ws.cell(row=i, column=2, value=val)
        v.fill = INPUT_FILL if not val else PatternFill("solid", start_color="E2EFDA")
        v.font = BASE_FONT
        v.alignment = LEFT
        v.border = BORDER
        ws.row_dimensions[i].height = 22

    # Dropdowns
    add_dv(ws, DV_CURRENCY,  f"B9")
    add_dv(ws, DV_DIRECTION, f"B10")
    add_dv(ws, DV_TIMEFRAME, f"B11")
    add_dv(ws, '"Trend-Following,Mean-Reversion,Breakout,Arbitrage,Event-Driven,Market-Making,Scalping,Carry"', "B12")

    # Legend
    legend_row = len(rows) + 5
    ws.cell(row=legend_row, column=1, value="LEGEND").font = BOLD_FONT
    ws.cell(row=legend_row + 1, column=1, value="Yellow = user input").fill = INPUT_FILL
    ws.cell(row=legend_row + 2, column=1, value="Green  = completed / sample").fill = PatternFill("solid", start_color="E2EFDA")
    ws.cell(row=legend_row + 3, column=1, value="Blue   = dropdown field").fill = PatternFill("solid", start_color="DDEBF7")

    set_widths(ws, [26, 30, 18, 18])
    ws.freeze_panes = "A3"


def build_instruments(ws, sample=False):
    headers = ["Symbol", "Asset Class", "Description", "Contract Size",
               "Pip Value (USD)", "Min Spread (pips)", "Typical Spread (pips)",
               "Swap Long (pts)", "Swap Short (pts)", "Margin %", "Commission (USD/lot)",
               "Trading Hours (GMT)", "Active?"]
    ncols = len(headers)
    title_row(ws, "INSTRUMENT UNIVERSE", ncols)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, ncols)

    sample_rows = [
        ["EURUSD", "FX Major", "Euro vs US Dollar",        100000, 10.00, 0.1, 0.8, -3.5,  0.5, 3.33, 3.5, "22:00-22:00",           "Yes"],
        ["GBPUSD", "FX Major", "Pound vs US Dollar",       100000, 10.00, 0.3, 1.2, -4.8,  1.2, 3.33, 3.5, "22:00-22:00",           "Yes"],
        ["USDJPY", "FX Major", "US Dollar vs Yen",         100000,  9.10, 0.2, 1.0, -2.1, -3.0, 3.33, 3.5, "22:00-22:00",           "Yes"],
        ["XAUUSD", "Commodity CFD", "Gold vs US Dollar",      100,  1.00, 1.5, 3.0, -8.5, -2.5, 5.00, 6.0, "22:01-20:59",           "Yes"],
        ["US30",   "Index CFD",    "Dow Jones 30",             1,  1.00, 1.0, 2.5, -1.2, -1.8, 2.00, 0.0, "22:01-20:59 (break 21:00-22:00)", "Yes"],
        ["GER40",  "Index CFD",    "DAX 40",                   1,  1.00, 0.8, 1.8, -0.9, -1.5, 2.00, 0.0, "06:15-20:00",           "Yes"],
        ["BTCUSD", "Crypto CFD",   "Bitcoin vs USD",           1,  1.00, 15,  35,  -15.0,-15.0,50.00, 0.0, "24/7",                  "No"],
    ]
    if sample:
        for r_idx, row in enumerate(sample_rows, start=3):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        end_row = 2 + len(sample_rows)
    else:
        end_row = 22  # 20 blank rows

    apply_band(ws, 3, end_row, ncols)

    # Number formats
    for r in range(3, end_row + 1):
        ws.cell(row=r, column=4).number_format = "#,##0"
        for col in (5, 6, 7, 8, 9, 11):
            ws.cell(row=r, column=col).number_format = "#,##0.00"
        ws.cell(row=r, column=10).number_format = "0.00%"

    # Dropdowns
    add_dv(ws, DV_ASSET_CLASS, f"B3:B{end_row}")
    add_dv(ws, DV_YESNO,       f"M3:M{end_row}")

    # Mark yellow for blank template input cells
    if not sample:
        for r in range(3, end_row + 1):
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).fill = INPUT_FILL if (r - 3) % 2 == 0 else PatternFill("solid", start_color="FFF8E1")

    set_widths(ws, [12, 14, 24, 14, 16, 16, 18, 14, 14, 12, 18, 30, 10])
    ws.freeze_panes = "B3"


def build_position_sizing(ws, sample=False):
    title_row(ws, "POSITION SIZING & RISK CALCULATOR", 4)
    # Inputs block
    inputs = [
        ("Account Equity (USD)",    100000  if sample else ""),
        ("Risk per Trade (%)",      0.01    if sample else ""),
        ("Max Concurrent Positions",4       if sample else ""),
        ("Max Daily Loss (%)",      0.03    if sample else ""),
        ("Max Drawdown Tolerance (%)", 0.15 if sample else ""),
        ("Leverage",                30      if sample else ""),
        ("Correlation Cap / Cluster", 2     if sample else ""),
    ]
    for i, (k, v) in enumerate(inputs, start=3):
        ws.cell(row=i, column=1, value=k).font = BOLD_FONT
        ws.cell(row=i, column=1).fill = SUBHEADER_FILL
        ws.cell(row=i, column=1).font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
        ws.cell(row=i, column=1).alignment = LEFT
        ws.cell(row=i, column=1).border = BORDER
        c = ws.cell(row=i, column=2, value=v)
        c.fill = INPUT_FILL if not sample else PatternFill("solid", start_color="E2EFDA")
        c.border = BORDER
        c.alignment = CENTER
        if "%" in k:
            c.number_format = "0.00%"
        elif k.startswith("Account") or k.startswith("Leverage"):
            c.number_format = "#,##0"

    # Sizing table (per symbol, dummy / formulas)
    tbl_start = 3 + len(inputs) + 2
    ws.cell(row=tbl_start, column=1, value="LOT SIZE MATRIX").font = BOLD_FONT
    ws.cell(row=tbl_start, column=1).fill = SUBHEADER_FILL
    ws.cell(row=tbl_start, column=1).font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)

    h = ["Symbol", "Stop Distance (pips)", "Pip Value (USD/lot)", "Risk $ (calc)", "Lot Size (calc)"]
    hdr_row = tbl_start + 1
    for i, v in enumerate(h, 1):
        ws.cell(row=hdr_row, column=i, value=v)
    style_header(ws, hdr_row, len(h))

    if sample:
        rows = [
            ("EURUSD", 30, 10.00),
            ("GBPUSD", 40, 10.00),
            ("USDJPY", 35,  9.10),
            ("XAUUSD", 80,  1.00),
            ("US30",  120,  1.00),
            ("GER40", 100,  1.00),
        ]
    else:
        rows = [("", "", "")] * 8

    for i, (sym, stop, pv) in enumerate(rows):
        r = hdr_row + 1 + i
        ws.cell(row=r, column=1, value=sym)
        ws.cell(row=r, column=2, value=stop).number_format = "#,##0"
        ws.cell(row=r, column=3, value=pv).number_format = "#,##0.00"
        # Risk $ = Equity * Risk %
        ws.cell(row=r, column=4, value="=$B$3*$B$4").number_format = "$#,##0.00"
        # Lot size = Risk $ / (Stop pips * Pip value)
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/(B{r}*C{r}),0)").number_format = "0.00"

    end = hdr_row + len(rows)
    apply_band(ws, hdr_row + 1, end, len(h))

    set_widths(ws, [30, 22, 22, 22, 22])
    ws.freeze_panes = "A3"


def build_session_hours(ws, sample=False):
    headers = ["Session", "Open (GMT)", "Close (GMT)", "Overlap With", "Best For", "Trade in This Session?"]
    ncols = len(headers)
    title_row(ws, "TRADING SESSIONS", ncols)
    for i, v in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=v)
    style_header(ws, 2, ncols)

    data = [
        ("Sydney",            "21:00", "06:00", "Tokyo",   "AUD, NZD pairs",       "No"),
        ("Tokyo",             "00:00", "09:00", "Sydney",  "JPY pairs, Asian indices", "No"),
        ("London",            "07:00", "16:00", "NY",      "EUR, GBP, CHF pairs",  "Yes"),
        ("New York",          "12:00", "21:00", "London",  "USD pairs, US indices","Yes"),
        ("London-NY Overlap", "12:00", "16:00", "—",       "Highest liquidity, best for breakouts", "Yes"),
    ]
    for r_idx, row in enumerate(data, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val if sample else (val if c_idx <= 3 else ""))

    end_row = 2 + len(data)
    apply_band(ws, 3, end_row, ncols)
    add_dv(ws, DV_YESNO, f"F3:F{end_row}")
    if not sample:
        for r in range(3, end_row + 1):
            for c in (4, 5, 6):
                ws.cell(row=r, column=c).fill = INPUT_FILL

    set_widths(ws, [22, 14, 14, 18, 38, 20])
    ws.freeze_panes = "A3"


def build_cost_model(ws, sample=False):
    headers = ["Symbol", "Spread Cost (USD/lot)", "Commission (USD/lot)",
               "Swap/Day (USD/lot)", "Slippage Assumption (pips)", "Total RT Cost (USD/lot)"]
    ncols = len(headers)
    title_row(ws, "COST MODEL PER INSTRUMENT", ncols)
    for i, v in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=v)
    style_header(ws, 2, ncols)

    if sample:
        rows = [
            ("EURUSD", 8.00, 3.5,  -3.50, 0.5),
            ("GBPUSD", 12.00, 3.5, -4.80, 0.5),
            ("USDJPY", 9.10, 3.5,  -2.10, 0.5),
            ("XAUUSD", 3.00, 6.0,  -8.50, 1.0),
            ("US30",   2.50, 0.0,  -1.20, 1.5),
            ("GER40",  1.80, 0.0,  -0.90, 1.5),
        ]
    else:
        rows = [("",)*5]*10

    for i, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            if c_idx >= 2 and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
        # Total RT cost = 2*spread + 2*commission + abs(slippage * pip value assumed = 10)
        ws.cell(row=i, column=6, value=f"=IFERROR(2*B{i}+2*C{i}+E{i}*10,0)").number_format = "#,##0.00"

    end = 2 + len(rows)
    apply_band(ws, 3, end, ncols)
    set_widths(ws, [14, 24, 24, 22, 28, 26])
    ws.freeze_panes = "B3"


def build_backtest_metrics(ws, sample=False):
    headers = ["Symbol", "Period", "# Trades", "Win Rate", "Avg R:R",
               "Sharpe", "Sortino", "Max DD", "CAGR", "Profit Factor", "Net PnL (USD)"]
    ncols = len(headers)
    title_row(ws, "BACKTEST METRICS (TRADER-PROVIDED)", ncols)
    for i, v in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=v)
    style_header(ws, 2, ncols)

    if sample:
        rows = [
            ("EURUSD", "2020-01 to 2025-12", 412, 0.52, 1.8, 1.42, 1.98, 0.082, 0.145, 1.63, 42350),
            ("GBPUSD", "2020-01 to 2025-12", 398, 0.49, 2.0, 1.28, 1.71, 0.094, 0.131, 1.55, 38720),
            ("USDJPY", "2020-01 to 2025-12", 385, 0.54, 1.6, 1.35, 1.89, 0.076, 0.128, 1.58, 35110),
            ("XAUUSD", "2020-01 to 2025-12", 287, 0.51, 2.2, 1.67, 2.31, 0.108, 0.186, 1.82, 58900),
            ("US30",   "2020-01 to 2025-12", 246, 0.56, 1.5, 1.51, 2.05, 0.089, 0.163, 1.71, 48200),
            ("GER40",  "2020-01 to 2025-12", 231, 0.53, 1.7, 1.44, 1.92, 0.097, 0.154, 1.66, 44100),
            ("PORTFOLIO", "2020-01 to 2025-12", 1959, 0.525, 1.8, 2.18, 2.94, 0.071, 0.162, 1.68, 267380),
        ]
    else:
        rows = [("",)*11]*10

    for i, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c_idx, value=val)
            if c_idx == 4 or c_idx == 8 or c_idx == 9:
                cell.number_format = "0.0%"
            elif c_idx in (5, 6, 7, 10):
                cell.number_format = "0.00"
            elif c_idx == 11:
                cell.number_format = "$#,##0"
            elif c_idx == 3:
                cell.number_format = "#,##0"

    end = 2 + len(rows)
    apply_band(ws, 3, end, ncols)

    # Highlight portfolio row if sample
    if sample:
        for c in range(1, ncols + 1):
            ws.cell(row=end, column=c).font = Font(name=FONT_NAME, bold=True, size=10)
            ws.cell(row=end, column=c).fill = PatternFill("solid", start_color="FFE699")

    set_widths(ws, [14, 22, 10, 12, 10, 10, 10, 10, 10, 14, 16])
    ws.freeze_panes = "B3"


def add_confidential_strip(ws, ncols):
    """Insert CONFIDENTIAL strip at top of sheet (pushes existing content down by 3)."""
    ws.insert_rows(1, amount=3)
    # Row 1: logo anchor spacer (row height)
    ws.row_dimensions[1].height = 50
    # Row 2: CONFIDENTIAL banner
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(ncols, 4))
    c = ws.cell(row=2, column=1, value="CONFIDENTIAL — REY CAPITAL | SMART INVESTMENTS")
    c.fill = CONFIDENTIAL_FILL
    c.font = CONFIDENTIAL_FONT
    c.alignment = CENTER
    ws.row_dimensions[2].height = 20
    # Row 3: spacer
    ws.row_dimensions[3].height = 6


def build_workbook(path, sample=False):
    wb = Workbook()
    # Remove default
    wb.remove(wb.active)

    sheets = [
        ("Cover",            build_cover,           4),
        ("Instruments",      build_instruments,     13),
        ("PositionSizing",   build_position_sizing, 5),
        ("SessionHours",     build_session_hours,   6),
        ("CostModel",        build_cost_model,      6),
        ("BacktestMetrics",  build_backtest_metrics,11),
    ]
    for name, builder, ncols in sheets:
        ws = wb.create_sheet(name)
        builder(ws, sample=sample)
        add_confidential_strip(ws, ncols)
        add_logo(ws, anchor="A1")
        add_confidential_header(ws)
        # shift freeze_panes down by 3 rows
        if ws.freeze_panes:
            fp = ws.freeze_panes
            # parse like "B3" -> "B6"
            col_letters = ''.join(ch for ch in fp if ch.isalpha())
            row_num = int(''.join(ch for ch in fp if ch.isdigit()))
            ws.freeze_panes = f"{col_letters}{row_num + 3}"

    # Workbook-level document properties
    wb.properties.title = "CFD Strategy Intake — Rey Capital"
    wb.properties.creator = "Rey Capital"
    wb.properties.company = "Rey Capital"
    wb.properties.keywords = "CONFIDENTIAL, CFD, Strategy, Rey Capital, DHRUVA"
    wb.properties.description = "Confidential — Rey Capital internal strategy intake document."

    wb.save(path)
    print(f"Saved: {path}")


if __name__ == "__main__":
    build_workbook(f"{OUT_DIR}/CFD_Strategy_Params_v1.xlsx", sample=False)
    build_workbook(f"{OUT_DIR}/CFD_Strategy_Params_SAMPLE_EMA_9_21_Crossover.xlsx", sample=True)
