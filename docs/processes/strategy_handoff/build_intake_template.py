"""Build the Strategy Intake & Evaluation Excel template (Rey Capital Automation Team).

Gate 0 + Gate 1 scoring workbook that manual traders fill before a strategy
is accepted for automation. Auto-scores and flags reject/accept.

Run:
    python build_intake_template.py
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Brand ─────────────────────────────────────────────────────────────────
BRAND_BLUE = "004AAC"
BRAND_LIGHT = "BDD4F5"
SOFT_BG = "F4F7FC"
WHITE = "FFFFFF"
DARK_TEXT = "1A1A1A"
MUTED_TEXT = "6B7280"
GREEN = "16A34A"
RED = "DC2626"
AMBER = "F59E0B"

FONT_NAME = "Calibri"

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parents[2]
LOGO = REPO_ROOT / "monitoring" / "dashboards" / "mt5_pnl_dashboard" / "src" / "TradeAtlas" / "Assets" / "ReyCapital_Logo.png"
OUTPUT = HERE / "Strategy_Intake_Evaluation_Template.xlsx"

# ── Reusable styles ───────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin", color="BFCBDB"),
    right=Side(style="thin", color="BFCBDB"),
    top=Side(style="thin", color="BFCBDB"),
    bottom=Side(style="thin", color="BFCBDB"),
)

HEADER_FILL = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
LABEL_FILL = PatternFill(start_color=SOFT_BG, end_color=SOFT_BG, fill_type="solid")
LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True, color=BRAND_BLUE)
VALUE_FONT = Font(name=FONT_NAME, size=10, color=DARK_TEXT)
MUTED_FONT = Font(name=FONT_NAME, size=9, italic=True, color=MUTED_TEXT)
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=BRAND_BLUE)
SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color=BRAND_BLUE)
SECTION_FILL = PatternFill(start_color=BRAND_LIGHT, end_color=BRAND_LIGHT, fill_type="solid")
PASS_FONT = Font(name=FONT_NAME, size=11, bold=True, color=GREEN)
FAIL_FONT = Font(name=FONT_NAME, size=11, bold=True, color=RED)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(vertical="center")


def style_cell(cell, *, font=VALUE_FONT, fill=None, alignment=LEFT_ALIGN, border=THIN_BORDER):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border


def write_section_header(ws, row, text, col_start=1, col_end=6):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    style_cell(cell, font=SECTION_FONT, fill=SECTION_FILL, alignment=LEFT_ALIGN)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = SECTION_FILL


def write_kv_row(ws, row, label, placeholder="", col_label=1, col_value=2, merge_to=None):
    lc = ws.cell(row=row, column=col_label, value=label)
    style_cell(lc, font=LABEL_FONT, fill=LABEL_FILL)
    vc = ws.cell(row=row, column=col_value, value=placeholder)
    style_cell(vc, font=MUTED_FONT)
    if merge_to and merge_to > col_value:
        ws.merge_cells(start_row=row, start_column=col_value, end_row=row, end_column=merge_to)
        for c in range(col_value, merge_to + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER


def write_table_header(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)


def write_table_row(ws, row, values, col_start=1, fonts=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        f = fonts[i] if fonts and i < len(fonts) else VALUE_FONT
        style_cell(cell, font=f, alignment=CENTER_ALIGN if i > 0 else LEFT_ALIGN)


def add_logo(ws, cell_ref="A1"):
    if LOGO.exists():
        img = XlImage(str(LOGO))
        img.width = 180
        img.height = 50
        ws.add_image(img, cell_ref)


# ── TAB 1: Strategy Intake Form ──────────────────────────────────────────

def build_intake_tab(wb: Workbook):
    ws = wb.active
    ws.title = "1. Strategy Intake"
    ws.sheet_properties.tabColor = BRAND_BLUE

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    row = 1
    # Logo & title
    add_logo(ws, "A1")
    ws.row_dimensions[1].height = 45
    ws.merge_cells("C1:F1")
    tc = ws.cell(row=1, column=3, value="Rey Capital Automation Team")
    style_cell(tc, font=TITLE_FONT)
    row = 2
    ws.merge_cells("C2:F2")
    sc = ws.cell(row=2, column=3, value="Strategy Intake & Evaluation Form")
    style_cell(sc, font=Font(name=FONT_NAME, size=13, bold=True, color=MUTED_TEXT))

    row = 4
    # ── Section: Strategy Identity ────────────────────────────────────
    write_section_header(ws, row, "STRATEGY IDENTITY")
    row += 1

    fields = [
        ("Strategy Name", "<unique short name>"),
        ("Author / Submitter", "<manual trader name>"),
        ("Submission Date", "<YYYY-MM-DD>"),
        ("Strategy Type", "<Trend Following / Mean Reversion / Momentum / Breakout / Scalping>"),
        ("Markets / Instruments", "<e.g. XAUUSD, NQ, NIFTY, EURUSD>"),
        ("Timeframe(s)", "<e.g. 5M chart, 1H confirmation>"),
        ("Trading Sessions", "<e.g. London 08:00-12:00, NY 09:30-11:30>"),
        ("Trade Direction", "<Long Only / Short Only / Both>"),
        ("Short Description of Edge", "<one paragraph: what market inefficiency does this exploit?>"),
    ]
    for label, placeholder in fields:
        write_kv_row(ws, row, label, placeholder, merge_to=6)
        row += 1

    row += 1
    # ── Section: Indicators & Configuration ───────────────────────────
    write_section_header(ws, row, "INDICATORS & CONFIGURATION")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    nc = ws.cell(row=row, column=1, value="List every indicator the strategy depends on. Anything not listed will NOT be implemented.")
    style_cell(nc, font=MUTED_FONT)
    row += 1

    ind_headers = ["#", "Indicator Name", "Parameters / Settings", "Applied To (TF / Price)", "Source / Library", "Purpose in Strategy"]
    write_table_header(ws, row, ind_headers)
    row += 1
    for i in range(1, 9):
        write_table_row(ws, row, [i, "", "", "", "", ""])
        row += 1

    row += 1
    # ── Section: Entry Logic ──────────────────────────────────────────
    write_section_header(ws, row, "ENTRY LOGIC (Step-by-Step)")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    nc = ws.cell(row=row, column=1, value="Write deterministic conditions only. No discretionary words (\"usually\", \"sometimes\", \"I feel\"). Each step must be codeable.")
    style_cell(nc, font=MUTED_FONT)
    row += 1

    entry_headers = ["Step #", "Condition Type", "Condition Description", "Indicator / Value Used", "Timeframe", "Required (Y/N)"]
    write_table_header(ws, row, entry_headers)
    row += 1
    condition_types = ["Signal,Filter,Confirmation,Confluence,Timing"
    ]
    dv_ctype = DataValidation(type="list", formula1='"Signal,Filter,Confirmation,Confluence,Timing"', allow_blank=True)
    dv_ctype.error = "Pick from: Signal, Filter, Confirmation, Confluence, Timing"
    dv_ctype.errorTitle = "Invalid condition type"
    ws.add_data_validation(dv_ctype)

    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_yn)

    entry_start = row
    for i in range(1, 11):
        write_table_row(ws, row, [i, "", "", "", "", ""])
        dv_ctype.add(ws.cell(row=row, column=2))
        dv_yn.add(ws.cell(row=row, column=6))
        row += 1

    row += 1
    # ── Section: Exit Logic ───────────────────────────────────────────
    write_section_header(ws, row, "EXIT LOGIC")
    row += 1

    exit_fields = [
        ("Stop Loss Type", "<Fixed pips / ATR-based / Structure-based / % equity>"),
        ("Stop Loss Value", "<e.g. 20 pips, 1.5x ATR, below swing low>"),
        ("Take Profit Type", "<Fixed R:R / Indicator signal / Trailing / Partial TP>"),
        ("Take Profit Value", "<e.g. 2R, opposite signal, 50% at 1R + trail rest>"),
        ("Trailing Stop Logic", "<e.g. trail by ATR(14) after 1R reached>"),
        ("Time-Based Exit", "<e.g. close at session end, max hold 4 hours>"),
        ("Partial Close Rules", "<e.g. close 50% at TP1, move SL to BE, trail rest>"),
        ("Break-Even Rule", "<e.g. move SL to entry after 1R profit>"),
    ]
    for label, placeholder in exit_fields:
        write_kv_row(ws, row, label, placeholder, merge_to=6)
        row += 1

    row += 1
    # ── Section: Filters & Confluences ────────────────────────────────
    write_section_header(ws, row, "FILTERS & CONFLUENCES")
    row += 1

    filter_fields = [
        ("Higher Timeframe Bias", "<e.g. only long if Daily EMA200 is rising>"),
        ("Session Filter", "<e.g. trade only during London/NY overlap>"),
        ("Volatility Filter", "<e.g. ATR(14) > 10 pips, avoid low-vol days>"),
        ("News Avoidance", "<e.g. no trades 30 min before/after high-impact news>"),
        ("Spread Filter", "<e.g. skip if spread > 2 pips>"),
        ("Day-of-Week Filter", "<e.g. no trades on Friday after 15:00>"),
        ("Correlation Filter", "<e.g. no two long positions on correlated pairs>"),
        ("Other Filters", "<any additional filters>"),
    ]
    for label, placeholder in filter_fields:
        write_kv_row(ws, row, label, placeholder, merge_to=6)
        row += 1

    row += 1
    # ── Section: Risk & Money Management ──────────────────────────────
    write_section_header(ws, row, "RISK & MONEY MANAGEMENT")
    row += 1

    risk_fields = [
        ("Risk Per Trade", "<e.g. 0.5% of equity, fixed 1 lot>"),
        ("Position Sizing Rule", "<Fixed Lot / % Equity / Volatility-Based / Kelly>"),
        ("Max Concurrent Positions", "<e.g. 2>"),
        ("Max Daily Loss Limit", "<e.g. 2% equity, hard cutoff>"),
        ("Max Weekly Loss Limit", "<e.g. 5% equity>"),
        ("Max Drawdown Stop", "<e.g. 8% — pause and review>"),
        ("Leverage Cap", "<e.g. 1:10>"),
        ("Scaling In/Out Rules", "<e.g. add 50% at confirmation, or N/A>"),
    ]
    for label, placeholder in risk_fields:
        write_kv_row(ws, row, label, placeholder, merge_to=6)
        row += 1

    row += 1
    # ── Section: Operational Details ──────────────────────────────────
    write_section_header(ws, row, "OPERATIONAL DETAILS")
    row += 1

    ops_fields = [
        ("Target Broker(s)", "<e.g. ICMarkets, Zerodha, IBKR>"),
        ("Platform", "<MT5 / TradingView / Freqtrade / Custom>"),
        ("Data Feed", "<Broker feed / External / TradingView>"),
        ("Expected Spread", "<e.g. ≤ 1.2 pips>"),
        ("Expected Slippage", "<e.g. 0.5 pip / 1 tick>"),
        ("Commission Model", "<per lot / per side / spread-included>"),
        ("Latency Tolerance", "<e.g. ≤ 200ms / not latency-sensitive>"),
        ("Required Platform Features", "<hedging, OCO, partial close, etc.>"),
    ]
    for label, placeholder in ops_fields:
        write_kv_row(ws, row, label, placeholder, merge_to=6)
        row += 1

    row += 1
    # ── Section: Worked Example ───────────────────────────────────────
    write_section_header(ws, row, "WORKED EXAMPLE TRADE (Mandatory)")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    nc = ws.cell(row=row, column=1, value="Describe one complete trade step-by-step. Attach annotated chart screenshot separately.")
    style_cell(nc, font=MUTED_FONT)
    row += 1

    example_fields = [
        ("Date & Time of Trade", ""),
        ("Instrument & Timeframe", ""),
        ("Market Context / Bias", ""),
        ("Step 1: Setup Identified", "<what did you see?>"),
        ("Step 2: Entry Trigger", "<exact trigger and entry price>"),
        ("Step 3: SL & TP Placed", "<levels and reasoning>"),
        ("Step 4: Trade Management", "<any adjustments during trade>"),
        ("Step 5: Exit & Result", "<exit reason, price, P&L in R>"),
        ("Chart Screenshot Attached?", "<Yes / No — attach separately>"),
    ]
    for label, placeholder in example_fields:
        write_kv_row(ws, row, label, placeholder, merge_to=6)
        row += 1

    return ws


# ── TAB 2: Backtest Metrics ───────────────────────────────────────────────

def build_backtest_tab(wb: Workbook):
    ws = wb.create_sheet("2. Backtest Metrics")
    ws.sheet_properties.tabColor = BRAND_BLUE

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    row = 1
    add_logo(ws, "A1")
    ws.row_dimensions[1].height = 45
    ws.merge_cells("C1:E1")
    tc = ws.cell(row=1, column=3, value="Backtest Performance Report")
    style_cell(tc, font=TITLE_FONT)

    row = 3
    write_section_header(ws, row, "BACKTEST CONTEXT", col_end=5)
    row += 1
    ctx_fields = [
        ("Backtest Platform", "<MT5 Strategy Tester / TradingView / Python / Excel>"),
        ("Data Source", "<broker history / TradingView / external provider>"),
        ("Backtest Start Date", "<YYYY-MM-DD>"),
        ("Backtest End Date", "<YYYY-MM-DD>"),
        ("Total Backtest Duration", "<e.g. 14 months>"),
        ("In-Sample Period", "<e.g. first 8 months>"),
        ("Out-of-Sample Period", "<e.g. last 6 months>"),
        ("Starting Capital", "<e.g. $10,000>"),
        ("Commission & Spread Included?", "<Yes / No>"),
        ("Slippage Modeled?", "<Yes / No — specify model>"),
    ]
    for label, placeholder in ctx_fields:
        write_kv_row(ws, row, label, placeholder, merge_to=5)
        row += 1

    row += 1
    write_section_header(ws, row, "PERFORMANCE METRICS vs. MINIMUM BAR", col_end=5)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    nc = ws.cell(row=row, column=1, value="Strategy must PASS ALL minimum thresholds. Any single FAIL = strategy is NOT accepted for automation.")
    style_cell(nc, font=Font(name=FONT_NAME, size=10, bold=True, color=RED))
    row += 1

    metric_headers = ["Metric", "Minimum Required", "Your Value", "Pass / Fail", "Notes"]
    write_table_header(ws, row, metric_headers)
    metrics_row_start = row + 1
    row += 1

    # Metrics with minimum thresholds
    metrics = [
        ("Total Trades", "≥ 100", "", ""),
        ("Backtest Duration", "≥ 12 months", "", ""),
        ("Out-of-Sample Period", "≥ 3 months", "", ""),
        ("Win Rate (%)", "≥ 40%", "", ""),
        ("Profit Factor", "≥ 1.5", "", ""),
        ("Sharpe Ratio (annualized)", "≥ 1.0", "", ""),
        ("Max Drawdown (%)", "≤ 15%", "", ""),
        ("Max Consecutive Losses", "≤ 10", "", ""),
        ("Avg Risk:Reward Ratio", "≥ 1.2", "", ""),
        ("Expectancy per Trade ($)", "> 0", "", ""),
        ("Recovery Factor", "≥ 2.0", "", ""),
        ("Calmar Ratio", "≥ 1.0", "", ""),
        ("Avg Trade Duration", "Document only", "", ""),
        ("Largest Single Loss", "Document only", "", ""),
        ("Largest Single Win", "Document only", "", ""),
        ("Longest Winning Streak", "Document only", "", ""),
        ("Longest Losing Streak", "Document only", "", ""),
        ("Monthly Return Std Dev", "Document only", "", ""),
    ]

    dv_pf = DataValidation(type="list", formula1='"PASS,FAIL,N/A"', allow_blank=True)
    dv_pf.error = "Select PASS, FAIL, or N/A"
    ws.add_data_validation(dv_pf)

    for metric_name, minimum, value, notes in metrics:
        write_table_row(ws, row, [metric_name, minimum, value, "", notes])
        dv_pf.add(ws.cell(row=row, column=4))
        row += 1

    row += 1
    write_section_header(ws, row, "EQUITY CURVE SUMMARY", col_end=5)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    nc = ws.cell(row=row, column=1, value="Attach full equity curve chart, monthly P&L breakdown, and drawdown chart as separate files.")
    style_cell(nc, font=MUTED_FONT)
    row += 1

    eq_fields = [
        ("Equity Curve Attached?", "<Yes / No>"),
        ("Monthly P&L Breakdown Attached?", "<Yes / No>"),
        ("Drawdown Chart Attached?", "<Yes / No>"),
        ("Trade List Export Attached?", "<Yes / No — CSV or Excel>"),
        ("Walk-Forward Report Attached?", "<Yes / No>"),
    ]
    for label, placeholder in eq_fields:
        write_kv_row(ws, row, label, placeholder, merge_to=5)
        row += 1

    row += 1
    write_section_header(ws, row, "MARKET REGIME PERFORMANCE", col_end=5)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    nc = ws.cell(row=row, column=1, value="How does the strategy perform across different market conditions? (must cover at least 2 regimes)")
    style_cell(nc, font=MUTED_FONT)
    row += 1

    regime_headers = ["Market Regime", "Period / Dates", "Trades", "Profit Factor", "Notes"]
    write_table_header(ws, row, regime_headers)
    row += 1
    regimes = ["Trending (Up)", "Trending (Down)", "Ranging / Sideways", "High Volatility", "Low Volatility"]
    for regime in regimes:
        write_table_row(ws, row, [regime, "", "", "", ""])
        row += 1

    return ws


# ── TAB 3: Scoring & Decision ────────────────────────────────────────────

def build_scoring_tab(wb: Workbook):
    ws = wb.create_sheet("3. Scoring & Decision")
    ws.sheet_properties.tabColor = BRAND_BLUE

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30

    row = 1
    add_logo(ws, "A1")
    ws.row_dimensions[1].height = 45
    ws.merge_cells("C1:E1")
    tc = ws.cell(row=1, column=3, value="Strategy Evaluation Scorecard")
    style_cell(tc, font=TITLE_FONT)

    # ── Auto-Reject Checklist ─────────────────────────────────────────
    row = 3
    write_section_header(ws, row, "AUTO-REJECT CHECKLIST (Any single YES = Immediate Reject)", col_end=5)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    nc = ws.cell(row=row, column=1, value="If ANY item below is YES, the strategy is automatically rejected. Do not proceed to scoring.")
    style_cell(nc, font=Font(name=FONT_NAME, size=10, bold=True, color=RED))
    row += 1

    reject_headers = ["Rejection Criterion", "Yes / No", "Auto-Reject?", "Reviewer Notes"]
    write_table_header(ws, row, reject_headers[:4])
    reject_start = row + 1
    row += 1

    dv_yn = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    ws.add_data_validation(dv_yn)

    reject_items = [
        "Profit Factor < 1.3",
        "Backtest duration < 12 months",
        "Fewer than 80 trades in backtest",
        "No stop loss defined",
        "Max drawdown > 20%",
        "Rules contain discretionary language (\"usually\", \"I feel\", \"depends\")",
        "No out-of-sample period tested",
        "Strategy works in only one market regime",
        "Submitter cannot explain WHY the edge works",
        "No worked example trade provided",
        "Backtest does not include commission & spread",
        "Over-optimized (in-sample >> out-of-sample by > 40%)",
    ]

    for item in reject_items:
        c1 = ws.cell(row=row, column=1, value=item)
        style_cell(c1)
        c2 = ws.cell(row=row, column=2, value="")
        style_cell(c2, alignment=CENTER_ALIGN)
        dv_yn.add(c2)
        # Formula: if B=YES then "REJECT" else "OK"
        c3 = ws.cell(row=row, column=3)
        c3.value = f'=IF(B{row}="YES","⛔ REJECT","✅ OK")'
        style_cell(c3, alignment=CENTER_ALIGN)
        c4 = ws.cell(row=row, column=4, value="")
        style_cell(c4)
        row += 1

    reject_end = row - 1
    row += 1

    # Auto-reject verdict
    verdict_cell = ws.cell(row=row, column=1, value="AUTO-REJECT VERDICT")
    style_cell(verdict_cell, font=Font(name=FONT_NAME, size=12, bold=True, color=DARK_TEXT))
    formula_cell = ws.cell(row=row, column=2)
    formula_cell.value = f'=IF(COUNTIF(C{reject_start}:C{reject_end},"⛔ REJECT")>0,"⛔ REJECTED","✅ PROCEED TO SCORING")'
    style_cell(formula_cell, font=Font(name=FONT_NAME, size=12, bold=True, color=DARK_TEXT), alignment=CENTER_ALIGN)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)

    row += 2
    # ── Weighted Scoring ──────────────────────────────────────────────
    write_section_header(ws, row, "WEIGHTED SCORING (Only if Auto-Reject = PROCEED)", col_end=5)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    nc = ws.cell(row=row, column=1, value="Rate each category 1-10. The weighted score determines accept/review/reject. Target: ≥ 70/100.")
    style_cell(nc, font=MUTED_FONT)
    row += 1

    score_headers = ["Evaluation Category", "Weight", "Score (1-10)", "Weighted Score", "Reviewer Comments"]
    write_table_header(ws, row, score_headers)
    score_start = row + 1
    row += 1

    dv_score = DataValidation(type="whole", operator="between", formula1="1", formula2="10", allow_blank=True)
    dv_score.error = "Score must be between 1 and 10"
    dv_score.errorTitle = "Invalid Score"
    ws.add_data_validation(dv_score)

    scoring_items = [
        ("Strategy Logic Clarity (rules are unambiguous, codeable)", 15),
        ("Backtest Quality (duration, sample size, OOS testing)", 15),
        ("Risk Management Completeness (SL, position sizing, limits)", 15),
        ("Profit Factor & Expectancy", 12),
        ("Drawdown & Recovery (max DD, recovery factor)", 10),
        ("Market Regime Robustness (works in 2+ regimes)", 10),
        ("Edge Explainability (WHY it works is clear)", 8),
        ("Operational Feasibility (broker, latency, spread realistic)", 5),
        ("Documentation Quality (worked example, charts, trade list)", 5),
        ("Parameter Sensitivity (not over-optimized)", 5),
    ]

    for item, weight in scoring_items:
        c1 = ws.cell(row=row, column=1, value=item)
        style_cell(c1)
        c2 = ws.cell(row=row, column=2, value=weight)
        style_cell(c2, alignment=CENTER_ALIGN)
        c3 = ws.cell(row=row, column=3, value="")
        style_cell(c3, alignment=CENTER_ALIGN)
        dv_score.add(c3)
        c4 = ws.cell(row=row, column=4)
        c4.value = f"=IF(C{row}<>\"\",B{row}*C{row}/10,0)"
        style_cell(c4, alignment=CENTER_ALIGN)
        c5 = ws.cell(row=row, column=5, value="")
        style_cell(c5)
        row += 1

    score_end = row - 1
    row += 1

    # Totals row
    c1 = ws.cell(row=row, column=1, value="TOTAL WEIGHTED SCORE")
    style_cell(c1, font=Font(name=FONT_NAME, size=12, bold=True, color=BRAND_BLUE))
    c2 = ws.cell(row=row, column=2, value=f"=SUM(B{score_start}:B{score_end})")
    style_cell(c2, font=Font(name=FONT_NAME, size=12, bold=True, color=BRAND_BLUE), alignment=CENTER_ALIGN)
    c4 = ws.cell(row=row, column=4)
    c4.value = f"=SUM(D{score_start}:D{score_end})"
    style_cell(c4, font=Font(name=FONT_NAME, size=14, bold=True, color=BRAND_BLUE), alignment=CENTER_ALIGN)
    c5 = ws.cell(row=row, column=5, value="out of 100")
    style_cell(c5, font=LABEL_FONT)

    row += 2
    # ── Final Verdict ─────────────────────────────────────────────────
    write_section_header(ws, row, "FINAL DECISION", col_end=5)
    row += 1

    score_total_row = row - 2  # row where total is

    c1 = ws.cell(row=row, column=1, value="Decision")
    style_cell(c1, font=Font(name=FONT_NAME, size=14, bold=True, color=DARK_TEXT))
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    dc = ws.cell(row=row, column=2)
    dc.value = f'=IF(COUNTIF(C{reject_start}:C{reject_end},"⛔ REJECT")>0,"⛔ REJECTED — Fix auto-reject items first",IF(D{score_total_row}>=70,"✅ ACCEPTED — Proceed to full handoff",IF(D{score_total_row}>=50,"⚠️ CONDITIONAL — Needs improvement, resubmit","❌ REJECTED — Score too low")))'
    style_cell(dc, font=Font(name=FONT_NAME, size=14, bold=True, color=DARK_TEXT), alignment=CENTER_ALIGN)

    row += 2
    # Decision scale reference
    scale = [
        ("70 – 100", "✅ ACCEPTED", "Strategy proceeds to full handoff & automation"),
        ("50 – 69", "⚠️ CONDITIONAL", "Submitter must address weak areas and resubmit"),
        ("0 – 49", "❌ REJECTED", "Strategy does not meet minimum standards"),
    ]
    write_table_header(ws, row, ["Score Range", "Decision", "Action Required"])
    row += 1
    for score_range, decision, action in scale:
        write_table_row(ws, row, [score_range, decision, action])
        row += 1

    row += 2
    # ── Sign-off ──────────────────────────────────────────────────────
    write_section_header(ws, row, "REVIEW SIGN-OFF", col_end=5)
    row += 1

    signoff_headers = ["Role", "Name", "Date", "Decision", "Signature"]
    write_table_header(ws, row, signoff_headers)
    row += 1

    dv_decision = DataValidation(type="list", formula1='"ACCEPT,CONDITIONAL,REJECT"', allow_blank=True)
    ws.add_data_validation(dv_decision)

    roles = ["Manual Team Lead (Submitter)", "Automation Team Lead (Reviewer)", "Risk Reviewer"]
    for role in roles:
        write_table_row(ws, row, [role, "", "", "", ""])
        dv_decision.add(ws.cell(row=row, column=4))
        row += 1

    row += 2
    c1 = ws.cell(row=row, column=1, value="Additional Comments / Conditions for Resubmission:")
    style_cell(c1, font=LABEL_FONT)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 4, end_column=5)
    c = ws.cell(row=row, column=1, value="")
    style_cell(c, alignment=Alignment(wrap_text=True, vertical="top"))
    for r in range(row, row + 5):
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN_BORDER

    return ws


# ── TAB 4: Reference Guide ───────────────────────────────────────────────

def build_reference_tab(wb: Workbook):
    ws = wb.create_sheet("4. Reference Guide")
    ws.sheet_properties.tabColor = BRAND_BLUE

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 55

    row = 1
    add_logo(ws, "A1")
    ws.row_dimensions[1].height = 45
    ws.merge_cells("B1:C1")
    tc = ws.cell(row=1, column=2, value="Reference Guide & Metric Definitions")
    style_cell(tc, font=TITLE_FONT)

    row = 3
    write_section_header(ws, row, "METRIC DEFINITIONS & ACCEPTABLE RANGES", col_end=3)
    row += 1

    write_table_header(ws, row, ["Metric", "Acceptable Range", "What It Means"])
    row += 1

    definitions = [
        ("Profit Factor", "≥ 1.5 (ideal: ≥ 1.75)", "Gross profit / Gross loss. Below 1.0 = losing strategy. 1.5+ shows a meaningful edge."),
        ("Sharpe Ratio", "≥ 1.0 (ideal: ≥ 1.5)", "Risk-adjusted return. Higher = better return per unit of risk. Below 0.5 is generally unacceptable."),
        ("Max Drawdown", "≤ 15% (hard limit: 20%)", "Largest peak-to-trough decline. Indicates worst-case pain. Must be survivable."),
        ("Win Rate", "≥ 40% (context-dependent)", "% of trades that are profitable. Must be evaluated alongside R:R ratio."),
        ("Risk:Reward Ratio", "≥ 1.2 (avg)", "Average win size / Average loss size. Low win rate needs high R:R to compensate."),
        ("Expectancy", "> $0 per trade", "(Win% × Avg Win) - (Loss% × Avg Loss). Must be positive = edge exists."),
        ("Recovery Factor", "≥ 2.0", "Net profit / Max drawdown. Measures how quickly the strategy recovers from losses."),
        ("Calmar Ratio", "≥ 1.0", "Annualized return / Max drawdown. Similar to Sharpe but uses drawdown instead of volatility."),
        ("Sample Size", "≥ 100 trades (ideal: 300+)", "Too few trades = results may be luck, not edge. Statistical significance requires volume."),
        ("Backtest Duration", "≥ 12 months (ideal: 2-5 yrs)", "Must span multiple market regimes. Short backtests capture one condition only."),
        ("Out-of-Sample %", "≥ 30% of total period", "Data the strategy was NOT optimized on. Tests if the edge generalizes."),
        ("t-Statistic", "≥ 2.0", "Statistical confidence that returns ≠ 0. Below 2.0 = edge may be random noise."),
        ("Max Consecutive Losses", "≤ 10", "Longest losing streak. Tests psychological and capital survivability."),
    ]
    for metric, acceptable, meaning in definitions:
        c1 = ws.cell(row=row, column=1, value=metric)
        style_cell(c1, font=LABEL_FONT, fill=LABEL_FILL)
        c2 = ws.cell(row=row, column=2, value=acceptable)
        style_cell(c2, alignment=CENTER_ALIGN)
        c3 = ws.cell(row=row, column=3, value=meaning)
        style_cell(c3, alignment=WRAP_ALIGN)
        ws.row_dimensions[row].height = 35
        row += 1

    row += 1
    write_section_header(ws, row, "STRATEGY EVALUATION PROCESS FLOW", col_end=3)
    row += 1

    process_steps = [
        ("Gate 0: Intake Form", "Manual trader", "Fill this Excel workbook (Tab 1 + Tab 2). Attach backtest reports, equity curve, trade list, and chart screenshots."),
        ("Gate 1: Auto-Reject Screen", "Automation team lead", "Review Tab 3 auto-reject checklist. If any item flags YES → return to submitter with feedback."),
        ("Gate 2: Scoring Review", "Automation team lead + Risk", "Score each category 1-10 in Tab 3. Total ≥ 70 = accepted. 50-69 = conditional resubmit. <50 = rejected."),
        ("Gate 3: Sign-Off", "Both team leads + Risk", "If accepted, all three reviewers sign off. Strategy moves to the full handoff Word document."),
        ("Gate 4: Full Handoff Doc", "Manual trader", "Complete the detailed Strategy Handoff & Go-Live Acceptance template (Word document)."),
        ("Gate 5: Implementation", "Automation team", "Code the strategy. Re-backtest to verify parity with manual results. Unit test all logic."),
        ("Gate 6: Paper Trading", "Automation team", "Run on demo/paper for 1-3 months. Results must match backtest within 1 std deviation."),
        ("Gate 7: Pilot Capital", "Both teams + Risk", "Deploy with 25% of target capital. Ramp: 25% → 50% → 75% → 100% over 3-6 months."),
        ("Gate 8: Full Deployment", "Head of Trading", "Final approval for 100% capital allocation. Ongoing monitoring begins."),
    ]

    write_table_header(ws, row, ["Stage", "Responsible", "Description"])
    row += 1
    for stage, responsible, desc in process_steps:
        c1 = ws.cell(row=row, column=1, value=stage)
        style_cell(c1, font=LABEL_FONT, fill=LABEL_FILL)
        c2 = ws.cell(row=row, column=2, value=responsible)
        style_cell(c2, alignment=CENTER_ALIGN)
        c3 = ws.cell(row=row, column=3, value=desc)
        style_cell(c3, alignment=WRAP_ALIGN)
        ws.row_dimensions[row].height = 40
        row += 1

    row += 1
    write_section_header(ws, row, "RECOMMENDED READING", col_end=3)
    row += 1

    books = [
        ("Advances in Financial Machine Learning", "Marcos Lopez de Prado", "Walk-forward validation, avoiding overfitting, combinatorial purged cross-validation"),
        ("Quantitative Trading", "Ernest Chan", "Practical strategy pipeline for smaller operations, backtesting methodology"),
        ("Systematic Trading", "Robert Carver", "Position sizing frameworks, capital allocation, risk management"),
    ]
    write_table_header(ws, row, ["Book", "Author", "Key Topics"])
    row += 1
    for book, author, topics in books:
        write_table_row(ws, row, [book, author, topics])
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    fc = ws.cell(row=row, column=1, value="Rey Capital Automation Team  •  Internal Use Only  •  Template v1.0")
    style_cell(fc, font=Font(name=FONT_NAME, size=9, italic=True, color=MUTED_TEXT), alignment=CENTER_ALIGN)

    return ws


# ── Build ─────────────────────────────────────────────────────────────────

def build() -> Path:
    wb = Workbook()

    build_intake_tab(wb)
    build_backtest_tab(wb)
    build_scoring_tab(wb)
    build_reference_tab(wb)

    # Print settings for all sheets
    for ws in wb.worksheets:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.gridLines = True

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
